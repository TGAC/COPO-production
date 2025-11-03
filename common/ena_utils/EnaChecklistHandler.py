import requests
from lxml import etree as ET
from common.dal.copo_da import EnaChecklist
from common.utils.logger import Logger
import pandas as pd
import os
import io
from django.conf import settings
from openpyxl.utils.cell import get_column_letter
from common.utils.helpers import (
    get_datetime,
    get_not_deleted_flag,
    get_env,
    notify_ena_object_status,
    describe_regex,
)
from django_tools.middlewares import ThreadLocal
import inspect
import math
from common.schema_versions.lookup import dtol_lookups as lookup
from common.validators.ena_validators import ena_checklist_validators  as required_validators
from common.validators.validator import Validator
import json
from common.dal.profile_da import Profile
from src.apps.copo_core.models import ProfileType

l = Logger()

class ChecklistHandler:
    def __init__(self):
        self.pass_word = get_env('WEBIN_USER_PASSWORD')
        self.user_token = get_env('WEBIN_USER').split("@")[0]
        self.headers = {'Accept': 'application/xml' }

    def _loadCheckList(self, url):
        with requests.Session() as session:    
            session.auth = (self.user_token, self.pass_word) 
            try:
                response = session.get(url,headers=self.headers)
                return response.text
            except Exception as e:
                l.exception(e)
                return ""            

    def _loadChecklist_local(self, file_location):
        '''
        This is a local file for testing purpose
        '''
        with open(file_location, 'r') as file:
            xmlstr = file.read()
            return xmlstr
    '''
    <CHECKLIST_SET>
    <CHECKLIST accession="ERT000028" checklistType="Sequence">
    <IDENTIFIERS>
        <PRIMARY_ID>ERT000028</PRIMARY_ID>
    </IDENTIFIERS>
    <DESCRIPTOR>
        <LABEL>Single Viral CDS</LABEL>
        <NAME>Single Viral CDS</NAME>
        <DESCRIPTION>For complete or partial single coding sequence (CDS) from a viral gene. Please do not use for peptides processed from polyproteins or proviral sequences, as these are all annotated differently.</DESCRIPTION>
        <AUTHORITY>ENA</AUTHORITY>
        <FIELD_GROUP restrictionType="Any number or none of the fields">
            <NAME>Mandatory Fields and Questions</NAME>
            <FIELD>
                <LABEL>VMOLTYPE</LABEL>
                <NAME>Molecule Type</NAME>
                <DESCRIPTION>Type of in vivo molecule sequenced. Taken from the INSDC controlled vocabulary. Example: Genomic DNA, Genomic RNA, viral cRNA.</DESCRIPTION>
                <FIELD_TYPE>
                <TEXT_CHOICE_FIELD>
                    <TEXT_VALUE>
                        <VALUE>genomic DNA</VALUE>
                    </TEXT_VALUE>
                    <TEXT_VALUE>
                        <VALUE>genomic RNA</VALUE>
                    </TEXT_VALUE>
                    <TEXT_VALUE>
                        <VALUE>viral cRNA</VALUE>
                    </TEXT_VALUE>
                </TEXT_CHOICE_FIELD>
                </FIELD_TYPE>
                <MANDATORY>mandatory</MANDATORY>
                <MULTIPLICITY>single</MULTIPLICITY>
            </FIELD>
            <FIELD>
                <LABEL>ORGANISM</LABEL>
                <NAME>Organism</NAME>
                <DESCRIPTION>Full name of virus (ICTV-approved or otherwise), NCBI taxid, BioSample accession, SRA sample accession, or sample alias. Influenza, Norovirus, Sapovirus and HIV have special nomenclature. Please contact us if you are unsure. Example: Raspberry bushy dwarf virus, Influenza A virus (A/chicken/Germany/1949(H10N7)), HIV-1 M:F_CHU51.</DESCRIPTION>
                <FIELD_TYPE>
                <TAXON_FIELD/>
                </FIELD_TYPE>
                <MANDATORY>mandatory</MANDATORY>
                <MULTIPLICITY>single</MULTIPLICITY>
            </FIELD>
            <FIELD>
                <LABEL>GENE</LABEL>
                <NAME>Gene</NAME>
                <DESCRIPTION>Symbol of the gene corresponding to a sequence region. Example: RdRp, CP, ORF1.</DESCRIPTION>
                <FIELD_TYPE>
                <TEXT_FIELD/>
                </FIELD_TYPE>
                <MANDATORY>mandatory</MANDATORY>
                <MULTIPLICITY>single</MULTIPLICITY>
            </FIELD>
            <FIELD>
                <LABEL>SVCGRTABLE</LABEL>
                <NAME>Translation table</NAME>
                <DESCRIPTION>Translation table for this virus. Chose between standard (table 1) and mitovirus codes (table 4.). Example: 1, 4.</DESCRIPTION>
                <FIELD_TYPE>
                <TEXT_FIELD/>
                </FIELD_TYPE>
                <MANDATORY>mandatory</MANDATORY>
                <MULTIPLICITY>single</MULTIPLICITY>
            </FIELD>
            <FIELD>
                <LABEL>5PARTIAL</LABEL>
                <NAME>Partial at 5' ? (yes/no)</NAME>
                <DESCRIPTION>For an incomplete CDS with the start codon upstream of the submitted sequence.</DESCRIPTION>
                <UNITS>
                  <UNIT>DD</UNIT>
                </UNITS>
                <FIELD_TYPE>
                <TEXT_CHOICE_FIELD>
                    <TEXT_VALUE>
                        <VALUE>yes</VALUE>
                    </TEXT_VALUE>
                    <TEXT_VALUE>
                        <VALUE>no</VALUE>
                    </TEXT_VALUE>
                </TEXT_CHOICE_FIELD>
                </FIELD_TYPE>
                <MANDATORY>mandatory</MANDATORY>
                <MULTIPLICITY>single</MULTIPLICITY>
            </FIELD>
            <FIELD>
                <LABEL>5CDS</LABEL>
                <NAME>5' CDS location</NAME>
                <DESCRIPTION>Start of the coding region relative to the submitted sequence. For a full length CDS this is the position of the first base of the start codon.</DESCRIPTION>
                <FIELD_TYPE>
                <TEXT_FIELD>
                    <REGEX_VALUE>\d+</REGEX_VALUE>
                </TEXT_FIELD> 
                </FIELD_TYPE>
                <MANDATORY>mandatory</MANDATORY>
                <MULTIPLICITY>single</MULTIPLICITY>
            </FIELD> 
            </FIELD_GROUP>
        </DESCRIPTOR>
        </CHECKLIST>
        </CHECKLIST_SET>
    '''

    def _parseCheckList(self, xmlstr):
        xml = xmlstr.encode('utf-8')
        parser = ET.XMLParser(ns_clean=True, recover=True, encoding='utf-8')
        checklist_set = []
        dt = get_datetime()
        root = ET.fromstring(xml, parser=parser)
        # checklist_ids = settings.BARCODING_CHECKLIST
        for checklist_elm in root.findall("./CHECKLIST"):
            checklist = {}

            primary_id = checklist_elm.find("./IDENTIFIERS/PRIMARY_ID").text.strip() 
            ena_checklist_id = checklist_elm.find("./IDENTIFIERS/ENA_CHECKLIST_ID")
            if ena_checklist_id is None:
                checklist['ena_checklist_id'] = primary_id
            else:
                checklist['ena_checklist_id'] = ena_checklist_id.text.strip()

            skip = settings.ENA_CHECKLIST_CONFIG.get(primary_id, dict()).get( "skip", list() )
            # if primary_id not in checklist_ids:
            #    continue
            # checklist_ids.remove(primary_id)
            checklist['primary_id'] = primary_id
            checklist['name'] = checklist_elm.find("./DESCRIPTOR/NAME").text.strip()
            checklist['description'] = checklist_elm.find("./DESCRIPTOR/DESCRIPTION").text.strip()
            checklist['fields'] = {}
            for field_elm in checklist_elm.findall("./DESCRIPTOR/FIELD_GROUP/FIELD"):
                try:
                    field = {}
                    key = field_elm.find("./NAME")
                    if key is None:
                        l.debug(f"{primary_id} Field does not have a name")
                        key = field_elm.find("./LABEL")
                        if key is None:
                            l.debug(f"{primary_id} Field does not have a label")
                            continue
                    key = key.text.strip()

                    if key in skip:
                        continue

                    label = field_elm.find("./LABEL")
                    if label is not None:
                        field['label'] = label.text.strip()
                    else: 
                        l.debug(f"{primary_id} Field {key} does not have a name")
                        field['label'] = key
                    field['name'] = key
                    desc = field_elm.find("./DESCRIPTION")
                    if desc is not None:
                        field['description'] = desc.text.strip()
                    field['mandatory'] = field_elm.find("./MANDATORY").text.strip()
                    field['multiplicity'] = field_elm.find("./MULTIPLICITY").text.strip()
                    synonym = field_elm.find("./SYNONYM")
                    if synonym is not None:
                        field['synonym'] = synonym.text.strip()

                    unit = field_elm.find("./UNITS/UNIT")
                    if unit is not None:
                        field['unit'] = unit.text.strip()
                    field['type'] = field_elm.find("./FIELD_TYPE")[0].tag
                    choice = field_elm.find("./FIELD_TYPE/TEXT_CHOICE_FIELD")
                    if choice is not None:
                        field['choice'] = []
                        for choice_elm in choice.findall("./TEXT_VALUE"):
                            field['choice'].append(choice_elm.find("./VALUE").text.strip())

                    regex = field_elm.find("./FIELD_TYPE/TEXT_FIELD/REGEX_VALUE")
                    if regex is not None:
                        field['regex'] = regex.text.strip()
                        field['regex_description'] = describe_regex(field['regex'])
                    ontology = field_elm.find("./FIELD_TYPE/ONTOLOGY")
                    if ontology is not None:
                        field['ontology'] = ontology.text.strip()
                        field['ontology_link'] = f"http://purl.obolibrary.org/obo/{str(field['ontology']).replace(':','_')}"
                    if key == 'language':
                        key = "language_code"
                    checklist['fields'][key] = field
                    if checklist['primary_id'].startswith("ERC"):
                        # add ORGANISM
                        field = {}
                        field['name'] = "Organism"
                        field['label'] = "Organism"
                        field['description'] = "Scientific Name"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "SCIENTIFIC_NAME_FIELD"
                        checklist['fields']["organism"] = field
                        # add SAMPLE
                        field = {}
                        field['name'] = "Sample"
                        field['label'] = "Sample"
                        field['description'] = "Sample ID or Name"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "TEXT_FIELD"
                        checklist['fields']["sample"] = field
                    elif checklist['primary_id'].startswith("ERT"):
                        # add SPECIMEN_ID
                        field = {}
                        field['name'] = "SPECIMEN_ID"
                        field['label'] = "SPECIMEN_ID"
                        field['description'] = "SPECIMENT_ID"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "TEXT_FIELD"
                        checklist['fields']["SPECIMEN_ID"] = field
                        # add TAXON_ID
                        field = {}
                        field['name'] = "TAXON_ID"
                        field['label'] = "TAXON_ID"
                        field['description'] = "TAXON_ID"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "TEXT_FIELD"
                        checklist['fields']["TAXON_ID"] = field
                except Exception as e:
                    l.exception(e)
                    l.debug(f"{primary_id} {key} Field does not have a name")
                    continue
            checklist["modified_date"] =  dt
            checklist["deleted"] = get_not_deleted_flag()
            checklist_set.append(checklist)
            # if len(checklist_ids) == 0:
            #    break
        return checklist_set

    def updateCheckList(self):
        urls = settings.ENA_CHECKLIST_URL
        urls.extend(settings.COPO_SAMPLE_CHECKLIST_URL)

        checklist_set = []

        for url in urls:
            xmlstr = self._loadCheckList(url)
            checklist_set.extend(self._parseCheckList(xmlstr))

        """
        xmlstr = self._loadChecklist_local("sample_checklist_dwc.xml")    
        checklist_set.extend(self._parseCheckList(xmlstr))

        xmlstr = self._loadChecklist_local("sample_checklist_faang.xml")
        checklist_set.extend(self._parseCheckList(xmlstr))
        """

        """
        reads = EnaChecklist().execute_query({"primary_id": "read", "deleted": get_not_deleted_flag()})
        read = None
        if reads:
           read = reads[0]
        """

        for checklist in checklist_set:
            # if checklist["primary_id"].startswith("ERC"):
            # read_fields = {key: value for key, value in read["fields"].items() if value.get("for_dtol", False) == False}
            # for key, value in read_fields.items():
            #    value.update({"read_field": True})
            # checklist["fields"].update(read_fields)
            EnaChecklist().get_collection_handle().find_one_and_update({"primary_id": checklist["primary_id"]},
                                                                            {"$set": checklist},
                                                                            upsert=True)
            # for checklist fields
            write_manifest(checklist, with_read=False)   

            """
            #for checklist feils + read fields 
            if checklist["primary_id"].startswith("ERC"): 
                fields = read["fields"]
                fields.update(checklist["fields"])
                checklist["fields"] = fields
                write_manifest(checklist, with_read=True)  
            """

class EnaCheckListSpreadsheet:
    def __init__(self, file, checklist_id, component, validators=[], with_read=True, with_sample=True):
        self.with_read = with_read
        self.req = ThreadLocal.get_current_request()
        self.profile_id = self.req.session.get("profile_id", None)
        profile = Profile().get_record(self.profile_id)
        profile_type = ProfileType.objects.get(type=profile["type"])

        self.checklist_id = checklist_id
        self.checklist = EnaChecklist().get_checklist(self.checklist_id, with_read=self.with_read, for_dtol=profile_type.is_dtol_profile, with_sample=with_sample)  
        self.data = None
        self.new_data = None
        self.component = component
        self.component_info = f"{self.component}_info"
        self.component_table = f"{self.component}_table"
        self.required_validators = []    
        self.symbiont_list = []
        self.validator_list = []
        # if a file is passed in, then this is the first time we have seen the spreadsheet,
        # if not then we are looking at creating samples having previously validated
        if file:
            self.file = file
        # else:
        #    self.sample_data = self.req.session.get( self.component_table, "")
        #    self.isupdate = self.req.session.get("isupdate", False)

        # create list of required validators
        required = dict(globals().items())["required_validators"]
        for element_name in dir(required):
            element = getattr(required, element_name)
            if inspect.isclass(element) and issubclass(element, Validator) and not element.__name__ == "Validator":
                self.required_validators.append(element)

        self.required_validators.extend(validators)

    def get_filenames_from_manifest(self):
        return list(self.data["File name"])

    def check_manifest_compatibility(self):
        ''' Checks whether the uploaded manifest's column set
            matches the expected checklist's columns.
        '''
        expected_columns = set(self.checklist['fields'].keys())
        uploaded_columns = set(self.data.columns)
        sheet_name = f"<strong>{self.checklist['primary_id']} {self.checklist['name']}</strong>"

        # Determine the number of expected columns that are missing
        missing = expected_columns - uploaded_columns

        # Calculate the match ratio
        total_expected = len(expected_columns)
        match_ratio = (total_expected - len(missing)) / max(total_expected, 1)

        # If almost nothing matches, show a single error message
        if match_ratio < 0.5:  # Threshold: less than 50% columns match
            raise Exception(
                'The uploaded manifest does not appear to match the expected checklist.<br><br>'
                'Please ensure that you are uploading the correct manifest for the selected checklist '
                f'(expected {len(expected_columns)} columns, found {len(uploaded_columns)} in sheet {sheet_name}).'
            )

    def loadManifest(self, m_format):

        if self.profile_id is not None:
            notify_ena_object_status(data={"profile_id": self.profile_id}, msg="Loading...", action="info",
                            html_id=self.component_info, checklist_id=self.checklist_id)

            try:
                # read excel and convert all to string
                if m_format == "xls":
                    self.data = pd.read_excel(self.file, keep_default_na=False,
                                                  na_values=lookup.NA_VALS)
                elif m_format == "csv":
                    self.data = pd.read_csv(self.file, keep_default_na=False,
                                                na_values=lookup.NA_VALS)
                else:
                    raise Exception("Unknown file format")
                if self.data.empty:
                    raise Exception("Empty file")
                self.data = self.data.loc[:, ~self.data.columns.str.contains('^Unnamed')]
                self.data = self.data.apply(lambda x: x.astype(str))
                self.data = self.data.apply(lambda x: x.str.strip())
                # self.data.columns = self.data.columns.str.replace(" ", "")

                new_column_name = { name : name.replace(" (optional)", "",-1).upper() for name in self.data.columns.values.tolist() }
                self.new_data = self.data.rename(columns=new_column_name)    

                new_column_name = { value["label"].upper() : key for key, value in self.checklist["fields"].items() }
                self.new_data.rename(columns=new_column_name, inplace=True)    

                self.check_manifest_compatibility()
            except Exception as e:
                # if error notify via web socket
                l.exception(e)
                notify_ena_object_status(data={"profile_id": self.profile_id}, msg="Unable to load file. " + str(e),
                                action="error",
                                html_id=self.component_info, checklist_id=self.checklist_id)
                return False
            return True

    def validate(self):
        flag = True
        errors = []
        warnings = []
        self.isupdate = False

        # validate for required fields
        for v in self.required_validators:
            try:
                errors, warnings, flag, self.isupdate = v(profile_id=self.profile_id, checklist=self.checklist,
                                                        data=self.new_data, fields=None,
                                                        errors=errors, warnings=warnings, flag=flag,
                                                        isupdate=self.isupdate).validate()
            except Exception as e:
                l.exception(e)
                error_message = str(e).replace("<", "").replace(">", "")

                flag = False
                errors.append(error_message)

                # notify_ena_object_status(data={"profile_id": self.profile_id}, msg="Server Error - " + error_message,
                #                action="info",
                #                html_id=self.component_info, checklist_id=self.checklist_id)

        # send warnings
        if warnings:
            l.log(",".join(warnings))
            notify_ena_object_status(data={"profile_id": self.profile_id},
                            msg="<br>".join(warnings),
                            action="warning",
                            html_id="warning_info2", checklist_id=self.checklist_id)
        # if flag is false, compile list of errors
        if not flag:
            errors = list(map(lambda x: "<li>" + x + "</li>", errors))
            errors = "".join(errors)
            l.log(errors)
            notify_ena_object_status(data={"profile_id": self.profile_id},
                            msg="<h4>" + self.file.name + "</h4><h2>Errors</h2><ol>" + errors + "</ol>",
                            action="error",
                            html_id=self.component_info, checklist_id=self.checklist_id)
            return False

        for column in self.new_data.columns:
            if column.startswith(Validator.PREFIX_4_NEW_FIELD):
                self.data[column.removeprefix(Validator.PREFIX_4_NEW_FIELD)] = self.new_data[column]

        # if we get here we have a valid spreadsheet
        notify_ena_object_status(data={"profile_id": self.profile_id}, msg="Spreadsheet is valid. Please click <b>Finish</b> to complete the upload.", 
            action="success", html_id=self.component_info)
        notify_ena_object_status(data={"profile_id": self.profile_id}, msg="", action="close", html_id="upload_controls", checklist_id=self.checklist_id)
        notify_ena_object_status(data={"profile_id": self.profile_id}, msg="", action="make_valid", html_id=self.component_info, checklist_id=self.checklist_id)

        return True

    def collect(self):
        # create table data to show to the frontend from parsed manifest
        tagged_seq_data = []
        headers = list()
        for col in list(self.data.columns):
            headers.append(col)
        tagged_seq_data.append(headers)
        for index, row in self.data.iterrows():
            r = list(row)
            for idx, x in enumerate(r):
                if x is math.nan:
                    r[idx] = ""
            tagged_seq_data.append(r)
        # store sample data in the session to be used to create mongo objects
        self.req.session[f"{self.component}_data"] = tagged_seq_data
        self.req.session["checklist_id"] = self.checklist_id

        notify_ena_object_status(data={"profile_id": self.profile_id}, msg=tagged_seq_data, action="make_table",
                        html_id=f"{self.component}_parse_table", checklist_id=self.checklist_id)


class ReadChecklistHandler:
    def __init__(self ):
        self.pass_word = get_env('WEBIN_USER_PASSWORD')
        self.user_token = get_env('WEBIN_USER').split("@")[0]
        self.headers = {'Accept': 'application/xml' }

    def _loadCheckList(self, url):
        with requests.Session() as session:    
            session.auth = (self.user_token, self.pass_word) 
            try:
                response = session.get(url,headers=self.headers)
                return response.text
            except Exception as e:
                l.exception(e)
                return ""         

    def _parseCheckList(self, jsonstr):
            dt = get_datetime()
            try:
                checklists_elm = json.loads(jsonstr)
                checklists = list()
                for checklist_elm in checklists_elm.get("fieldTypes",[]):
                    name = checklist_elm["name"]
                    if name == "fastq1":
                        checklist = dict()
                        checklist["primary_id"] = "read"
                        checklist["name"] = "read checklist"
                        checklist["description"] = checklist_elm["description"]
                        checklist["fields"] = {}
                        checklist["modified_date"] =  dt
                        checklist["deleted"] = get_not_deleted_flag()
                        for field_elm in checklist_elm["fields"]:

                            field = dict()
                            name = field_elm["name"]
                            if name.lower() in ["study"]:
                                continue

                            field["name"] = name
                            field["label"] = field_elm["label"]
                            field["description"] = field_elm["description"]
                            field["mandatory"] =  "mandatory" if field_elm.get("mandatory", False) else "optional"
                            field["type"] = "TEXT_FIELD"
                            if "value_choice" in field_elm:
                                field["type"] = "TEXT_CHOICE_FIELD"
                                field["choice"] = field_elm["value_choice"]
                            field["read_field"] = True
                            #don't want to have "sample" in the read checklist for DTOL profile
                            if field_elm["name"] == "sample":
                                field["for_dtol"] = False
                            checklist["fields"][field_elm["name"]] = field

                        """
                        field = {}
                        field['name'] = "Organism"
                        field['description'] = "Scientific Name"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "TEXT_FIELD"
                        #field["for_dtol"] = False
                        field["shown_when_no_sample"] = True
                        field["read_field"] = True
                        checklist['fields']["organism"] = field
                        """
                        field = {}
                        field['name'] = "biosampleAccession"
                        field['label'] = "biosampleAccession"
                        field['description'] = "Biosample Accession"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "BIOSAMPLEACCESSION_FIELD"
                        field["shown_when_no_sample"] = True
                        field["read_field"] = True
                        checklist['fields']["biosampleAccession"] = field  

                        field = {}
                        field['name'] = "taxon_id"
                        field['label'] = "taxon_id"
                        field['description'] = "taxon_id"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "BIOSAMPLEACCESSION_EXT_FIELD"
                        field["shown_when_no_sample"] = True
                        field["for_dtol"] = False
                        field["read_field"] = True
                        checklist['fields']["taxon_id"] = field  

                        #add SPECIMEN_ID
                        field = {}
                        field['name'] = "SPECIMEN_ID"
                        field['label'] = "SPECIMEN_ID"
                        field['description'] = "SPECIMENT_ID"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "BIOSAMPLEACCESSION_EXT_FIELD"
                        field["shown_when_no_sample"] = True
                        field["for_dtol"] = True
                        field["read_field"] = True
                        checklist['fields']["SPECIMEN_ID"] = field
                        
                        #add TAXON_ID
                        field = {}
                        field['name'] = "TAXON_ID"
                        field['label'] = "TAXON_ID"
                        field['description'] = "TAXON_ID"
                        field['mandatory'] = "mandatory"
                        field['multiplicity'] = "single"
                        field['type'] = "BIOSAMPLEACCESSION_EXT_FIELD"
                        field["shown_when_no_sample"] = True
                        field["for_dtol"] = True
                        field["read_field"] = True
                        checklist['fields']["TAXON_ID"] = field

                         
                        checklists.append(checklist)
                        break

                return checklists
            except Exception as e:
                l.exception(e)
                return []

    def updateCheckList(self):
        urls = [
            "https://www.ebi.ac.uk/ena/submit/report/checklists/getReadFields?format=xml"
        ]
        checklist_set = []
        for url in urls:
            jsonstr = self._loadCheckList(url)
            checklist_set.extend(self._parseCheckList(jsonstr))

        for checklist in checklist_set:

            EnaChecklist().get_collection_handle().find_one_and_update({"primary_id": checklist["primary_id"]},
                                                                            {"$set": checklist},
                                                                            upsert=True)        
            write_manifest(checklist, for_dtol=True, with_sample=False)
            write_manifest(checklist, for_dtol=False, with_sample=False)


def write_manifest(checklist, for_dtol=False, with_read=True, with_sample=True, samples=None, file_path=None):
    df = pd.DataFrame.from_dict(list(checklist["fields"].values()), orient='columns')

    """
    if for_dtol:
        df["for_dtol"] = df["for_dtol"].fillna(True)
        df = df.loc[df["for_dtol"] == True]
    else:
        if "for_dtol" in df.columns:
            df["for_dtol"] = df["for_dtol"].fillna(False)
            df = df.loc[df["for_dtol"] == False]

    if with_sample and with_read:
        df = df.loc[(df["shown_when_no_sample"] == False) | (df["shown_when_no_sample"].isnull())]
    """
    df.sort_values(by=['mandatory','name'], inplace=True)
    df.loc[df["mandatory"] == "mandatory" , "label"] = df["label"]
    df.loc[df["mandatory"] != "mandatory", "label"] = df["label"] + " (optional)"

    df1 = df.transpose()
    df1 = df1.loc[["label"]]
    df1.columns = df1.iloc[0]

    version = settings.MANIFEST_VERSION.get(checklist["primary_id"], str())
    if version:
        version = "_v" + version

    if samples is not None:
        sample_df = pd.DataFrame.from_records(samples)
        new_column_name = { key : field["label" ]+ (" (optional)" if  field["mandatory"] != 'mandatory' else "") for key, field in checklist["fields"].items() }
        sample_df.rename(columns=new_column_name, inplace=True)
        sample_df.drop(columns=sample_df.columns.difference(df1.columns), axis=1, inplace=True)

        #sample_df = sample_df.rename(columns={"name": "Sample"})
        df1 = pd.concat([df1, sample_df], axis=0, join="outer")
        df1 = df1.fillna("")

    if file_path is None:
        manifest_name = checklist["primary_id"]
        if with_read:
            if with_sample:
                manifest_name = manifest_name + "_read"

            if for_dtol:
                manifest_name = manifest_name + "_dtol"

        file_path = os.path.join(settings.MANIFEST_PATH, settings.MANIFEST_FILE_NAME.format(manifest_name, version)  )
 
    with pd.ExcelWriter(path=file_path, engine='xlsxwriter' ) as writer:  
        sheet_name = checklist["primary_id"] + " " + checklist["name"]
        sheet_name = sheet_name[:31]
        df1.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        data_validation_column_index = 0
        for field in checklist["fields"].values():
            name = field["label"] if field["mandatory"] == "mandatory"  else field["label"] + " (optional)"
            type = field.get("type","TEXT_FIELD")
            if name not in df1.columns:
                continue
            column_index = df1.columns.get_loc(name)
            column_length = len(name)
            cell_format = writer.book.add_format()
            if type.startswith("TEXT_"):
                cell_format.set_num_format('@')
            writer.sheets[sheet_name].set_column(column_index, column_index, column_length, cell_format)

            if type == "TEXT_CHOICE_FIELD" and "choice" in field:
                choice = field["choice"]
                column_letter = get_column_letter(column_index + 1)
                cell_start_end = '%s2:%s1048576' % (column_letter, column_letter)
                if len(choice) > 0:
                    source = ""
                    number_of_char_for_choice = sum([len(x) for x in choice])
                    if number_of_char_for_choice <= 255:
                        source = choice
                    else:
                        s = pd.Series(choice, name=field["label"])
                        s.to_frame().to_excel(writer, sheet_name="data_values", index=False, header=True, startrow=0, startcol=data_validation_column_index)
                        column_letter = get_column_letter(data_validation_column_index + 1)
                        column_length = max(s.astype(str).map(len).max(), len(field["name"]))
                        writer.sheets["data_values"].set_column(data_validation_column_index, data_validation_column_index, column_length)
                        source = "=%s!$%s$2:$%s$%s" % ("data_values", column_letter, column_letter, str(len(choice) + 1))
                        data_validation_column_index = data_validation_column_index + 1
                    writer.sheets[sheet_name].data_validation(cell_start_end,
                                                            {'validate': 'list',
                                                            'source': source})


        sheet_name = 'field_descriptions'           
        df.to_excel(writer, sheet_name=sheet_name)

        for column in df.columns:
            column_length = max(df[column].astype(str).map(len).max(), len(column))
            column_index = df.columns.get_loc(column)+1
            writer.sheets[sheet_name].set_column(column_index, column_index, column_length)    
