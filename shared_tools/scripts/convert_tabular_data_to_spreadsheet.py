'''
Purpose: To convert tabular data that is in string format to a 
         pandas DataFrame and save it as a spreadsheet file.
To run the script: $ python shared_tools/scripts/convert_tabular_data_to_spreadsheet.py

Alternatively, to execute the script via VSCode configuration, set the following in `launch.json` file:
    {
        "name": "Python: Convert Tabular Data to Spreadsheet",
        "type": "debugpy",
        "request": "launch",
        "program": "${workspaceFolder}/shared_tools/scripts/convert_tabular_data_to_spreadsheet.py",
        "console": "integratedTerminal",
        "env": {
        "PYTHONPATH": "${workspaceFolder}/lib:${PYTHONPATH}"
        },
        "args": [],
        "justMyCode": false,
        "pythonArgs": ["-Xfrozen_modules=off"]
    }
'''
import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

def convert_tabular_data_to_df(tabular_data):
    '''Convert tabular grid data into a cleaned pandas DataFrame.'''
    # Parse the string to remove borders and extract tabular data
    lines = tabular_data.strip().split('\n')

    # Remove border lines and any lines that consist only of '+' or '-' characters (separator lines)
    lines = [line for line in lines if not set(line.strip()) <= {'+', '-'}]

    # The first line is the columnheader
    header = lines[0].split('|')
    header = [col.strip() for col in header]

    # The rest are data rows
    data = [line.split('|') for line in lines[2:]]  # Skip the header and border

    # Clean up each column entry by stripping any leading/trailing spaces
    data = [[col.strip() for col in row] for row in data]

    # Create a DataFrame from the parsed data
    df = pd.DataFrame(data, columns=header)
    
     # Convert 'Sample count' column to numeric, forcing errors='coerce' to handle any non-numeric values
    if 'Sample count' in df.columns:
        df['Sample count'] = pd.to_numeric(df['Sample count'], errors='coerce')

    # Drop any fully empty columns
    df.dropna(axis=1, how='all', inplace=True)
    
    # Drop any fully empty rows (in case they exist)
    df.dropna(axis=0, how='all', inplace=True)
    
    return df

def generate_spreadsheet(df, file_path):
    '''Save the DataFrame to an Excel file with formatted headers and column adjustments.'''
    # Check if the file exists and remove it if it does
    if os.path.exists(file_path):
        os.remove(file_path)

    # Write the DataFrame to an Excel file
    df.to_excel(file_path, index=False, engine='openpyxl')
    
    # Load the workbook to format headers
    wb = load_workbook(file_path)
    ws = wb.active

    # Define border style
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Apply bold and centreed formatting to headers
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Format 'Sample count' column as a number (if it exists)
    if 'Sample count' in df.columns:
        col_index = df.columns.get_loc('Sample count') + 1  # Excel columns are 1-based index
        for cell in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=2):
            for c in cell:
                c.number_format = '#,##0'  # Format as a number with commas
    
    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding for better spacing
        
    # Save changes
    wb.save(file_path)

    print(f'Excel file \'{file_path}\' has been created in directory: {os.path.dirname(os.path.abspath(file_path))}')
    
if __name__ == '__main__':
    # NB: Replace the 'tabular_data' string with the actual tabular data you want to convert
    # Example usage
    tabular_data = '''
    +------------------+--------------+-------------------------+
    | Genomic profile  | Sample count | Owner email address     |
    +------------------+--------------+-------------------------+
    | Profile 1        | 150          | owner1@example.com      |
    | Profile 2        | 200          | owner2@example.com      |
    | Profile 3        | 120          | owner3@example.com      |
    +------------------+--------------+-------------------------+
    '''
    
    df = convert_tabular_data_to_df(tabular_data)
    
    # NB: Set the file path to the desired location
    file_path = 'output.xlsx'
    generate_spreadsheet(df, file_path)