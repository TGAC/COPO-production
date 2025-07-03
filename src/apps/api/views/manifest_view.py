import json
import jsonpath_rw_ext as jp
import os
import numpy as np
import operator
import pandas as pd
import re

from bson import ObjectId, json_util
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse, resolve
from io import BytesIO
from openpyxl.utils import cell
from rest_framework import status

import common.schema_versions.lookup.dtol_lookups as lkup
from common.dal.copo_da import DataFile, EnaChecklist
from common.dal.sample_da import Sample
from common.dal.profile_da import Profile
from common.ena_utils.EnaChecklistHandler import write_manifest as write_ena_manifest
from common.lookup import lookup as lk
from common.utils import helpers
from src.apps.copo_core.models import ProfileType
from src.apps.copo_single_cell_submission.utils.SingleCellSchemasHandler import (
    SingleCellSchemasHandler,
)
from src.apps.copo_single_cell_submission.utils.copo_single_cell import (
    SinglecellSchemas,
)
from src.apps.copo_core.templatetags.web_tags import get_blank_manifest_url, get_sop_url


def copo_manifests(request):
    context = {
        'manifest_rows': [
            {
                'prefix': 'copoSamples',
                'label_prefix': 'COPO',
                'label_suffix': 'samples',
                'title': 'Collaborative OPen Omics',
            },
            {
                'prefix': 'enaSamples',
                'label_prefix': 'ENA',
                'label_suffix': 'samples',
                'title': 'European Nucleotide Archive',
            },
            {
                'prefix': 'tolSamples',
                'label_prefix': 'ToL',
                'label_suffix': 'samples',
                'title': 'Tree of Life',
            },
            {
                'prefix': 'enaBarcode',
                'label_prefix': 'ENA',
                'label_suffix': 'tagged sequences or barcodes',
                'title': 'European Nucleotide Archive',
            },
            {
                'prefix': 'enaRead',
                'label_prefix': 'ENA',
                'label_suffix': 'reads',
                'title': 'European Nucleotide Archive',
            },
            {
                'prefix': 'singleCell',
                'label_prefix': 'Single-cell',
                'label_suffix': 'manifests',
                'title': '',
            },
        ]
    }

    return render(
        request,
        'manifests.html',
        context,
    )


def build_manifest_item(
    x,
    schema_name,
    manifest_type,
    label=None,
    show_prefill_wizard=False,
    blank_manifest_url=None,
    sop_url=None,
    technology=None,
    version=None,
):
    primary_id = x.get('primary_id', '')
    name = (x.get('name') or '').strip()
    description = x.get('description', '')
    file_name = os.path.basename(blank_manifest_url)

    return {
        'manifest_id': primary_id,
        'type': manifest_type,
        'technology': technology,
        'name': name,
        'label': label or f'{primary_id}: {name}',
        'description': description,
        'blank_manifest_url': blank_manifest_url,
        'sop_url': sop_url,
        'show_prefill_wizard': show_prefill_wizard,
        'file_name': file_name,
        'version': version,
    }


def populate_manifest_table(request):
    manifests = []

    checklist_service = EnaChecklist()
    profile_service = ProfileType()
    singlecell_service = SinglecellSchemas()

    # COPO samples
    copo_checklists = checklist_service.get_general_sample_checklists_no_fields()
    copo_items = [
        build_manifest_item(
            x,
            schema_name='COPO_SAMPLES',
            manifest_type='copo_samples',
            label=x['name'],
            blank_manifest_url=get_blank_manifest_url(x['primary_id']),
        )
        for x in copo_checklists
        if x['primary_id'].startswith('COPO')
    ]
    copo_items.sort(key=lambda x: x['label'])
    manifests.extend(copo_items)

    # ENA samples
    ena_items = [
        build_manifest_item(
            x,
            schema_name='ENA_SAMPLES',
            manifest_type='ena_samples',
            blank_manifest_url=get_blank_manifest_url(x['primary_id']),
        )
        for x in copo_checklists
        if not x['primary_id'].startswith('COPO')
    ]
    ena_items.sort(key=lambda x: x['label'])
    manifests.extend(ena_items)

    # ENA barcode
    barcode_checklists = checklist_service.get_barcoding_checklists_no_fields()
    barcode_items = [
        build_manifest_item(
            x,
            schema_name='ENA_BARCODE',
            manifest_type='ena_barcode',
            label=f"{x.get('primary_id', '')}: {(x.get('name') or '').strip()} checklist",
            blank_manifest_url=get_blank_manifest_url(x['primary_id']),
        )
        for x in barcode_checklists
    ]
    manifests.extend(barcode_items)

    # ENA read
    read_checklists = checklist_service.get_sample_checklists_no_fields()
    read_items = [
        build_manifest_item(
            x,
            schema_name='general',
            manifest_type='ena_read',
            label=(x.get('name') or '')
            .title()
            .replace('Checklist', 'checklist')
            .strip(),
            blank_manifest_url=get_blank_manifest_url(x['primary_id']),
            technology=None,
        )
        for x in read_checklists
        if x['primary_id'] == 'read'
    ]
    manifests.extend(read_items)

    # ToL samples
    excluded = ['dtolenv']
    tol_items = []

    for profile in profile_service.get_profile_types():
        if profile.is_dtol_profile and profile.type not in excluded:
            profile_type = profile.type.upper()
            blank_url = get_blank_manifest_url(profile_type)
            tol_items.append(
                {
                    'manifest_id': None,
                    'type': 'tol_samples',
                    'technology': None,
                    'name': profile_type,
                    'label': profile.description,
                    'description': None,
                    'blank_manifest_url': blank_url,
                    'sop_url': get_sop_url(profile_type),
                    'show_prefill_wizard': True,
                    'file_name': os.path.basename(blank_url),
                    'version': settings.MANIFEST_VERSION.get(profile_type, ''),
                }
            )
    manifests.extend(tol_items)

    # Single-cell
    schema_name = 'COPO_SINGLE_CELL'
    single_cell_manifests = singlecell_service.get_checklists(
        schema_name=schema_name, checklist_id=''
    )
    for checklist_id, value in single_cell_manifests.items():
        version = settings.MANIFEST_VERSION.get(schema_name, '')
        blank_manifest_url = get_blank_manifest_url(schema_name, checklist_id)

        item = {
            'manifest_id': checklist_id,
            'type': 'single_cell',
            'technology': value.get('technology', ''),
            'name': f"{value.get('standard', '')}_{value.get('technology', '')}",
            'label': value.get('name', ''),
            'description': value.get('description', ''),
            'blank_manifest_url': blank_manifest_url,
            'sop_url': None,
            'show_prefill_wizard': False,
            'file_name': os.path.basename(blank_manifest_url),
            'version': version,
        }
        manifests.append(item)

    return JsonResponse(manifests, safe=False)


def get_manifest_fields(request):
    manifest_type = request.GET['manifestType'].lower()
    current_schema_version = ''

    # Get manifest version
    current_schema_version = settings.MANIFEST_VERSION.get(manifest_type.upper(), '')

    # Get sample fields
    s = helpers.json_to_pytype(
        lk.WIZARD_FILES['sample_details'], compatibility_mode=False
    )

    field_lst = jp.match(
        '$.properties[?(@.specifications[*] == "'
        + manifest_type
        + '"& @.manifest_version[*]=="'
        + current_schema_version
        + '")].versions['
        '0]',
        s,
    )

    # Get sample fields' order number
    order_num_lst = jp.match(
        '$.properties[?(@.specifications[*] == "'
        + manifest_type
        + '"& @.manifest_version[*]=="'
        + current_schema_version
        + '")].order',
        s,
    )

    # Get sample fields' MS Excel column letter
    excel_col_lst = jp.match(
        '$.properties[?(@.specifications[*] == "'
        + manifest_type
        + '"& @.manifest_version[*]=="'
        + current_schema_version
        + '")].excel_col',
        s,
    )

    # Get sample fields' colour
    colour_lst = jp.match(
        '$.properties[?(@.specifications[*] == "'
        + manifest_type
        + '"& @.manifest_version[*]=="'
        + current_schema_version
        + '")].colour',
        s,
    )

    # Combine the information in a tuple
    sample_fields_tuple = tuple(
        zip(order_num_lst, excel_col_lst, colour_lst, field_lst)
    )

    # Filter the list of tuples
    # Get a list of tuples with field names that only begin with an uppercase letter
    field_lst_filtered_tup = list(
        filter(lambda x: x[3].isupper() == True, sample_fields_tuple)
    )
    field_lst_filtered_tup = list(set(field_lst_filtered_tup))  # Remove duplicates

    # Get a list of tuples that have blank/no value for order number, Excel column letter and colour
    sample_fields_with_no_values = list(
        filter(
            lambda x: x[0] is None and x[1] in (None, '') and x[2] in (None, ''),
            field_lst_filtered_tup,
        )
    )

    # Get a list of tuples that have inputted values for order number (integer),
    # Excel column letter (string) and colour (string)
    sample_fields_with_values = list(
        filter(
            lambda x: x[0] is not None
            and isinstance(x[0], int)
            and x[1] not in (None, '')
            and isinstance(x[1], str)
            and x[2] not in (None, '')
            and isinstance(x[2], str),
            field_lst_filtered_tup,
        )
    )

    # Sort the list of tuples that have values by order number
    sample_fields_with_values.sort(
        key=operator.itemgetter(0)
    )  # NB: itemgetter is faster than lambda to sort

    # Append the list of tuples with no values to the list of sorted tuples that have inputted values
    for x in sample_fields_with_no_values:
        sample_fields_with_values.append(x)

    sample_fields = (
        sample_fields_with_values  # Reassign for clarity now that everything is done
    )

    return HttpResponse(json.dumps(sample_fields))


def get_common_value_dropdown_list(request):
    date_fields = [field for field in lkup.DTOL_RULES if 'DATE' in field]
    integer_fields = [
        field
        for field in lkup.DTOL_RULES
        if lkup.DTOL_RULES.get(field, '').get('human_readable', '') == 'integer'
        or lkup.DTOL_RULES.get(field, '').get('human_readable', '') == 'numeric'
        or 'digit number' in lkup.DTOL_RULES.get(field, '').get('human_readable', '')
    ]

    manifest_type = request.GET['manifestType'].lower()
    common_field = request.GET['commonField']
    common_value_dropdownlist = get_common_field_dropdown_list(
        common_field, manifest_type
    )

    return HttpResponse(
        json.dumps(
            {
                'dropdown_list': common_value_dropdownlist,
                'date_fields': date_fields,
                'integer_fields': integer_fields,
            }
        )
    )


def compute_manifest_file_name(manifest_type):
    manifest_type = manifest_type.upper()

    if 'ASG' in manifest_type:
        type = 'ASG'
    elif 'DTOLENV' in manifest_type or 'ENV' in manifest_type:
        type = 'DTOLENV'
    elif 'DTOL' in manifest_type:
        type = 'DTOL'
    elif 'ERGA' in manifest_type:
        type = 'ERGA'
    else:
        type = manifest_type

    if manifest_type not in settings.MANIFEST_VERSION:
        return ''

    # Optional version string from settings
    version = settings.MANIFEST_VERSION.get(type, '')
    version_str = f'_v{version}' if version else ''

    return settings.MANIFEST_FILE_NAME.format(type, version_str)


def get_manifest_file_name(request, manifest_type):
    file_name = compute_manifest_file_name(manifest_type)
    return JsonResponse({'file_name': file_name})


def prefill_manifest_template(request):
    manifest_type = json_util.loads(request.body)['manifestType']
    number_of_samples = int(json_util.loads(request.body)['rowCount'])  # Convert to int
    common_fields = json_util.loads(request.body)['commonFieldsList']
    common_values = json_util.loads(request.body)['commonValuesList']

    # Set the path to the blank manifest template based on the manifest type
    file_name = compute_manifest_file_name(manifest_type)

    manifest_template_path = settings.BASE_DIR + get_blank_manifest_url(
        manifest_type.upper()
    )

    # Check if the manifest template file exists
    if not os.path.exists(manifest_template_path):
        return JsonResponse(
            {'error': f'Manifest template file {file_name} not found.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    # Duplicate the common field value according to the number of samples desired
    row_values = [[i] * int(number_of_samples) for i in common_values]

    # Create an Excel file dataframe for the common fields and values using dictionary comprehension
    excel_data = {common_fields[i]: row_values[i] for i in range(len(common_fields))}

    common_field_values_dataframe = pd.DataFrame(excel_data)

    bytesIO = generate_manifest_template(
        manifest_type, manifest_template_path, common_field_values_dataframe
    )

    '''    
    # Get all worksheets from the blank manifest
    blank_manifest_dataframe = pd.read_excel(manifest_template_path, sheet_name=None, index_col=None)

    # Get Metadata Entry worksheet
    metadataEntry_worksheet = blank_manifest_dataframe['Metadata Entry']

    # Get Data Validation worksheet
    dataValidation_worksheet = blank_manifest_dataframe['Data Validation']

    # Get OrganismPartDefinitions worksheet
    organismPartDefinitions_worksheet = blank_manifest_dataframe['OrganismPartDefinitions']

    # Get only the column names from the blank manifest
    # Convert the list of column names into a dataframe
    metadataEntry_worksheet_dataframe = pd.DataFrame(columns=metadataEntry_worksheet.columns)

    # Remove Unnamed columns
    metadataEntry_worksheet_dataframe = metadataEntry_worksheet_dataframe.loc[:,
                                        ~metadataEntry_worksheet_dataframe.columns.str.startswith(
                                            'Unnamed')]
    # Remove NaNs columns
    metadataEntry_worksheet_dataframe.dropna(axis=0, how='all', inplace=True)

    # Get column names from the 'Data Validation' worksheet from the blank manifest
    dataValidation_worksheet_dataframe = pd.DataFrame(columns=dataValidation_worksheet.columns, index=[0])

    # Concatenate the common field and its common values with the respective column names
    # from the blank manifest template
    metadataEntry_worksheet_concatenation = pd.concat(
        [metadataEntry_worksheet_dataframe, common_field_values_dataframe],
        ignore_index=True)

    bytesIO = BytesIO()

    with pd.ExcelWriter(bytesIO, engine='xlsxwriter' ) as pandas_writer:  


        # Add Metadata Entry worksheet to the generated manifest
        # worksheet using data from the blank manifest worksheet
        metadataEntry_worksheet_concatenation.to_excel(pandas_writer, index=False, startrow=0, sheet_name='Metadata Entry')

        # Remove Unnamed columns
        dataValidation_worksheet_dataframe = dataValidation_worksheet_dataframe.loc[:,
                                            ~dataValidation_worksheet_dataframe.columns.str.startswith(
                                                'Unnamed')]
        # Remove NaNs columns
        dataValidation_worksheet_dataframe.dropna(axis=0, how='all', inplace=True)

        # Add Data Validation worksheet to the generated manifest
        # worksheet using data from the blank manifest worksheet
        dataValidation_worksheet.to_excel(pandas_writer, index=False, startrow=0, sheet_name='Data Validation')

        # Add OrganismPartDefinitions worksheet to the generated manifest
        # worksheet using datafrom the blank manifest worksheet
        organismPartDefinitions_worksheet.to_excel(pandas_writer, index=False, startrow=0,
                                                sheet_name='OrganismPartDefinitions')

        # Auto-adjust width of each column within the worksheet
        auto_adjust_excel_column_width(metadataEntry_worksheet_concatenation, pandas_writer, 'Metadata Entry')

        auto_adjust_excel_column_width(dataValidation_worksheet, pandas_writer, 'Data Validation')

        auto_adjust_excel_column_width(organismPartDefinitions_worksheet, pandas_writer, 'OrganismPartDefinitions')

        # Apply a dropdown list to the desired columns
        apply_dropdown_list(metadataEntry_worksheet_concatenation, pandas_writer, 'Metadata Entry',
                        metadataEntry_worksheet_dataframe, dataValidation_worksheet_dataframe, manifest_type)

        #pandas_writer.save() '''

    bytesIO.seek(0)
    excel_workbook = bytesIO.getvalue()

    response = HttpResponse(
        excel_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}'

    return response


def auto_adjust_excel_column_width(dataframe, pandas_writer, sheet_name):
    for column in dataframe:
        column_length = max(dataframe[column].astype(str).map(len).max(), len(column))
        column_index = dataframe.columns.get_loc(column)
        pandas_writer.sheets[sheet_name].set_column(
            column_index, column_index, column_length
        )


def apply_data_validation_to_column(
    column,
    metadataEntry_worksheet_dataframe,
    dataValidation_worksheet_dataframe,
    pandas_writer,
    sheet_name,
):
    # Metadata Entry worksheet
    metadataEntry_worksheet_column_index = (
        metadataEntry_worksheet_dataframe.columns.get_loc(column)
    )
    metadataEntry_worksheet_column_letter = cell.get_column_letter(
        metadataEntry_worksheet_column_index + 1
    )

    # Get first row to the last row in a column from the 'Metadata Entry' worksheet
    row_start_end = '%s2:%s1048576' % (
        metadataEntry_worksheet_column_letter,
        metadataEntry_worksheet_column_letter,
    )

    # Data validation worksheet
    dataValidation_worksheet_name = "'Data Validation'"
    dataValidation_worksheet_column_index = dataValidation_worksheet_dataframe.columns.get_loc(column)
    dataValidation_worksheet_column_letter = cell.get_column_letter(dataValidation_worksheet_column_index + 1)
    #organismPart_dataValidationColumn = '=%s!$%s$2:$%s$78'
    #tissueForBarcoding_dataValidationColumn = '=%s!$%s$2:$%s$79'
    #tissueForBiobanking_dataValidationColumn = '=%s!$%s$2:$%s$79'
    dataValidationColumn = '=%s!$%s$2:$%s1000'

    data_validation_column = dataValidationColumn % (
        dataValidation_worksheet_name, dataValidation_worksheet_column_letter,
        dataValidation_worksheet_column_letter)

    return pandas_writer.sheets[sheet_name].data_validation(
        row_start_end, {'validate': 'list', 'source': data_validation_column})

    """
    if "ORGANISM_PART" in column:
        # Get dropdownlist from the first to last row of the column from the "Data Validation" worksheet
        data_validation_column = organismPart_dataValidationColumn % (
            dataValidation_worksheet_name,
            dataValidation_worksheet_column_letter,
            dataValidation_worksheet_column_letter,
        )

        return pandas_writer.sheets[sheet_name].data_validation(
            row_start_end, {'validate': 'list', 'source': data_validation_column}
        )
    elif 'TISSUE_FOR_BARCODING' in column:
        # Get dropdownlist from the first to last row of the column from the 'Data Validation' worksheet
        data_validation_column = tissueForBarcoding_dataValidationColumn % (
            dataValidation_worksheet_name,
            dataValidation_worksheet_column_letter,
            dataValidation_worksheet_column_letter,
        )

        return pandas_writer.sheets[sheet_name].data_validation(
            row_start_end, {'validate': 'list', 'source': data_validation_column}
        )
    elif 'TISSUE_FOR_BIOBANKING' in column:

        # Get dropdownlist from the first to last row of the column from the 'Data Validation' worksheet
        data_validation_column = tissueForBiobanking_dataValidationColumn % (
            dataValidation_worksheet_name,
            dataValidation_worksheet_column_letter,
            dataValidation_worksheet_column_letter,
        )

        return pandas_writer.sheets[sheet_name].data_validation(row_start_end,
                                                                {'validate': 'list', 'source': data_validation_column})
    """

def apply_dropdown_list(
    dataframe,
    pandas_writer,
    sheet_name,
    metadataEntry_worksheet_dataframe,
    dataValidation_worksheet_dataframe,
    manifest_type,
):
    for column_name in dataframe:
        column_length = max(
            dataframe[column_name].astype(str).map(len).max(), len(column_name)
        )
        column_index = dataframe.columns.get_loc(column_name)

        # Check if sheet is 'Metadata Entry' and column name is present amongst
        # the fields that require a dropdownlist
        if column_name in lkup.DTOL_ENUMS:
            # Get MS Excel official column header letter
            # Indexing starts at 0 by default but in this case, it should start at 1 so increment by 1
            column_letter = cell.get_column_letter(column_index + 1)

            '''MS Excel and Pandas XlsxWriter have a 255 character limit on list/string validation
            therefore, not all items will be listed in a dropdown list
            => columns like 'ORGANISM_PART', 'TISSUE_FOR_BARCODING' and 'TISSUE_FOR_BIOBANKING'
            exceed the 255 character limit so the dropdown list for each of these columns will 
            be pulled from the respective column in the in the 'Data Validation' worksheet'''

            common_value_dropdownlist = get_common_field_dropdown_list(column_name, manifest_type)
            if common_value_dropdownlist:
                common_value_dropdownlist.sort()

                number_of_characters = sum(len(i) for i in common_value_dropdownlist)
                max_length_str = max(common_value_dropdownlist, key=len)
                column_length = max(len(max_length_str), len(column_name))

                if number_of_characters >= 255:
                    apply_data_validation_to_column(
                        column_name,
                        metadataEntry_worksheet_dataframe,
                        dataValidation_worksheet_dataframe,
                        pandas_writer,
                        sheet_name,
                    )
                else:
                    # Get first row to the last row in a column
                    row_start_end = '%s2:%s1048576' % (column_letter, column_letter)

                    pandas_writer.sheets[sheet_name].data_validation(row_start_end,
                                                                        {'validate': 'list',
                                                                        'source': common_value_dropdownlist})

        pandas_writer.sheets[sheet_name].set_column(column_index, column_index, column_length)

def get_common_field_dropdown_list(common_field, manifest_type):
    def get_dropdown_items():
        if fieldsBasedOnManifestType and common_field in fieldsBasedOnManifestType:
            # Get dropdown list based on the manifest type
            dropdownlist = lkup.DTOL_ENUMS[common_field][manifest_type.upper()]
        elif common_field == 'COLLECTION_LOCATION':
            # 'COLLECTION_LOCATION' column does not require a dropdownlist
            dropdownlist = []
        else:
            # Get dropdown list
            dropdownlist = lkup.DTOL_ENUMS.get(common_field, [])

        dropdownlist.sort()  # Sort the list in ascending order
        return dropdownlist

    fieldsBasedOnManifestType = [
        field
        for field in lkup.DTOL_ENUMS
        if not isinstance(lkup.DTOL_ENUMS.get(field, ''), list)
        and lkup.DTOL_ENUMS.get(field, '').get(manifest_type.upper(), '')
    ]

    common_value_dropdownlist = []

    # The 'common field' parameter can be a list or not because the function,
    # get_common_field_dropdown_list(common_field, manifest_type), is called in a
    # couple functions in this file, ajax_handlers.py for various purposes

    # Check if the 'common field' parameter is a list or not
    if isinstance(common_field, list):
        common_field_lst = common_field  # Reassign the variable for naming convention's sake, readability and clarity
        # Iterate through the list of common fields for each common field
        for common_field in common_field_lst:
            common_value_dropdownlist = get_dropdown_items()
    else:
        # 'Common field' parameter is not a list
        common_value_dropdownlist = get_dropdown_items()

    return common_value_dropdownlist


def validate_common_value(request):
    common_field = request.GET['commonField']
    common_value = request.GET['commonValue']
    isCommonValueValid = False
    error_message = ''

    if (
        common_field not in lkup.DTOL_RULES
        and common_field != 'COLLECTION_LOCATION'
        and common_field != 'ORIGINAL_FIELD_COLLECTION_LOCATION'
    ):
        isCommonValueValid = True

    elif (
        common_field == 'COLLECTION_LOCATION'
        or common_field == 'ORIGINAL_FIELD_COLLECTION_LOCATION'
    ):
        # Validate 'COLLECTION_LOCATION' or 'ORIGINAL_FIELD_COLLECTION_LOCATION' value
        country_value = common_value.split('|')[0].strip()
        location_2part = common_value.split('|')[1:]
        isCommonValueValid = True

        if (
            country_value.upper() not in lkup.DTOL_ENUMS[common_field]
            or not location_2part
            or country_value.upper() in lkup.DTOL_ENUMS[common_field]
            and not location_2part
        ):
            isCommonValueValid = False
            error_message = (
                f'a specific location ranging from a least location to a most '
                f'specific location separated by | character. e.g. “United Kingdom | East Anglia | '
                f'Norfolk | Norwich | University of East Anglia | UEA Broad”. '
                f'See a list of allowed Country entries at '
                f'https://www.ebi.ac.uk/ena/browser/view/ERC000053 '
            )

    else:
        field_regex = ''
        if (
            'strict_regex' in lkup.DTOL_RULES[common_field]
            and 'ena_regex' in lkup.DTOL_RULES[common_field]
        ):
            field_regex = lkup.DTOL_RULES[common_field].get('strict_regex', 'ena_regex')
        elif 'ena_regex' in lkup.DTOL_RULES[common_field]:
            field_regex = lkup.DTOL_RULES[common_field]['ena_regex']
        elif 'strict_regex' in lkup.DTOL_RULES[common_field]:
            field_regex = lkup.DTOL_RULES[common_field]['strict_regex']
        elif 'optional_regex' in lkup.DTOL_RULES[common_field]:
            #  'optional_regex' in lkup.DTOL_RULES[common_field]
            field_regex = lkup.DTOL_RULES[common_field]['optional_regex']

        if 'human_readable' in lkup.DTOL_RULES[common_field]:
            error_message = lkup.DTOL_RULES[common_field]['human_readable']

        pattern = re.compile(f'r{field_regex}')
        isCommonValueValid = bool(pattern.match(common_value))

    if isCommonValueValid:
        return HttpResponse(json.dumps({'response': isCommonValueValid}))
    else:
        return HttpResponse(
            json.dumps({'response': isCommonValueValid, 'error': error_message})
        )


@login_required
def download_manifest(request, manifest_id):
    manifest_type = None
    profile = None
    samples = Sample().get_all_records_columns(filter_by={'manifest_id': manifest_id})
    if samples is None:
        return HttpResponse(status=404, content='Manifest not found')
    else:
        manifest_type = samples[0].get('sample_type')
        profile = Profile().get_record(samples[0].get('profile_id'))

    # if not sampl['species_list'][0]['SYMBIONT'] or sampl['species_list'][0]['SYMBIONT'] == 'TARGET':
    # special handling for species_list
    for sample in samples:
        if sample.get('species_list', []):
            sample['SYMBIONT'] = sample['species_list'][0].get('SYMBIONT', 'TARGET')
        else:
            sample['SYMBIONT'] = 'TARGET'

    # Special handling for 'POP_GENOMICS' associated project type
    # The following code is no longer needed since 'RESEQUENCING' is now a valid value for
    # 'PURPOSE_OF_SPECIMEN' as manifest v2.5
    '''
     for associated_type in  profile.get('associated_type', []) :
         if associated_type.get('value', '') == 'POP_GENOMICS':
             for sample in samples:
                 if sample.get('PURPOSE_OF_SPECIMEN', []) == 'RESEQUENCING':
                     sample['PURPOSE_OF_SPECIMEN'] = 'SHORT_READ_SEQUENCING'
    '''

    # special handling for permit file
    sample_df = pd.DataFrame.from_records(samples)
    for prefix in lkup.PERMIT_COLUMN_NAMES_PREFIX:
        permit_required_column = f'{prefix}_REQUIRED'
        if permit_required_column in sample_df.columns:
            file_name_column = f'{prefix}_FILENAME'
            if file_name_column in sample_df.columns:
                sample_df[file_name_column] = np.where(
                    sample_df[permit_required_column] == 'Y',
                    sample_df[file_name_column].apply(
                        lambda x: x.rsplit('_', 1)[0] + '.pdf' if '_' in x else x
                    ),
                    'NOT_APPLICABLE',
                )
            else:
                sample_df[file_name_column] = np.where(
                    sample_df[permit_required_column] == 'Y', '', 'NOT_APPLICABLE'
                )

    # manifests_dir = os.path.join('static', 'assets', 'manifests')

    # Set the path to the blank manifest template based on the manifest type
    file_name = compute_manifest_file_name(manifest_type)

    manifest_template_path = os.path.join(settings.MANIFEST_PATH, file_name)

    bytesIO = generate_manifest_template(
        manifest_type, manifest_template_path, sample_df
    )

    bytesIO.seek(0)
    excel_workbook = bytesIO.getvalue()

    response = HttpResponse(
        excel_workbook,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}'

    return response


@login_required
def download_permits(request):
    # Get array from Ajax call
    sampleIDs = json.loads(request.POST.get('sample_ids', []))

    permit_file_names = Sample().get_permit_file_names_by_sample_id(sampleIDs)

    if permit_file_names is None:
        return HttpResponse(json.dumps(list()))  # No permits found

    # Get profile ID using one of the sample IDs i.e. the first sample
    sample = Sample().get_by_field('_id', [ObjectId(sampleIDs[0])])
    profile_id = sample[0].get('profile_id', '')

    sample_permits_directory = os.path.join(
        settings.MEDIA_ROOT, 'sample_permits', profile_id
    )

    if not os.path.exists(sample_permits_directory):
        return HttpResponse(json.dumps(list()))  # No permits found

    permit_files_urls = list()

    for file_path, dirs, files in os.walk(sample_permits_directory):
        for file_name in files:
            if file_name.endswith('.pdf'):
                if any(file_name == x for x in permit_file_names) or any(
                    file_name == x.replace(f"_{x.split('_')[-1]}", '.pdf')
                    for x in permit_file_names
                ):
                    url = f"{file_path.replace('/copo',str())}/{file_name}"
                    permit_files_urls.append(url)

    return HttpResponse(json.dumps(permit_files_urls))


@login_required
def view_images(request):
    # Get array from Ajax call
    specimen_ids = json.loads(request.POST.get('specimen_ids', []))
    image_file_names = DataFile().get_image_file_names_by_specimen_id(specimen_ids)

    return HttpResponse(json.dumps(image_file_names))


def generate_manifest_template(manifest_type, manifest_template_path, initial_data):

    # Get all worksheets from the blank manifest
    blank_manifest_dataframe = pd.read_excel(
        manifest_template_path, sheet_name=None, index_col=None
    )

    # Get Metadata Entry worksheet
    metadataEntry_worksheet = blank_manifest_dataframe['Metadata Entry']

    # Get Data Validation worksheet
    dataValidation_worksheet = blank_manifest_dataframe['Data Validation']

    # Get OrganismPartDefinitions worksheet
    organismPartDefinitions_worksheet = blank_manifest_dataframe[
        'OrganismPartDefinitions'
    ]

    # Get only the column names from the blank manifest
    # Convert the list of column names into a dataframe
    metadataEntry_worksheet_dataframe = pd.DataFrame(
        columns=metadataEntry_worksheet.columns
    )

    # Remove Unnamed columns
    metadataEntry_worksheet_dataframe = metadataEntry_worksheet_dataframe.loc[
        :, ~metadataEntry_worksheet_dataframe.columns.str.startswith('Unnamed')
    ]
    # Remove NaNs columns
    metadataEntry_worksheet_dataframe.dropna(axis=0, how='all', inplace=True)

    # Get column names from the 'Data Validation' worksheet from the blank manifest
    dataValidation_worksheet_dataframe = pd.DataFrame(
        columns=dataValidation_worksheet.columns, index=[0]
    )

    # Concatenate the common field and its common values with the respective column names
    # from the blank manifest template
    initial_data.drop(
        columns=initial_data.columns.difference(
            metadataEntry_worksheet_dataframe.columns
        ),
        axis=1,
        inplace=True,
    )
    metadataEntry_worksheet_concatenation = pd.concat(
        [metadataEntry_worksheet_dataframe, initial_data], ignore_index=True
    )

    bytesIO = BytesIO()

    with pd.ExcelWriter(bytesIO, engine='xlsxwriter') as pandas_writer:

        # Add Metadata Entry worksheet to the generated manifest
        # worksheet using data from the blank manifest worksheet
        metadataEntry_worksheet_concatenation.to_excel(
            pandas_writer, index=False, startrow=0, sheet_name='Metadata Entry'
        )

        # Remove Unnamed columns
        dataValidation_worksheet_dataframe = dataValidation_worksheet_dataframe.loc[
            :, ~dataValidation_worksheet_dataframe.columns.str.startswith('Unnamed')
        ]
        # Remove NaNs columns
        dataValidation_worksheet_dataframe.dropna(axis=0, how='all', inplace=True)

        # Add Data Validation worksheet to the generated manifest
        # worksheet using data from the blank manifest worksheet
        dataValidation_worksheet.to_excel(
            pandas_writer, index=False, startrow=0, sheet_name='Data Validation'
        )

        # Add OrganismPartDefinitions worksheet to the generated manifest
        # worksheet using datafrom the blank manifest worksheet
        organismPartDefinitions_worksheet.to_excel(
            pandas_writer, index=False, startrow=0, sheet_name='OrganismPartDefinitions'
        )

        # Auto-adjust width of each column within the worksheet
        #autoAdjustExcelColumnWidth(metadataEntry_worksheet_concatenation, pandas_writer, 'Metadata Entry')

        auto_adjust_excel_column_width(
            dataValidation_worksheet, pandas_writer, 'Data Validation'
        )

        auto_adjust_excel_column_width(
            organismPartDefinitions_worksheet, pandas_writer, 'OrganismPartDefinitions'
        )

        # Apply a dropdown list to the desired columns
        apply_dropdown_list(
            metadataEntry_worksheet_concatenation,
            pandas_writer,
            'Metadata Entry',
            metadataEntry_worksheet_dataframe,
            dataValidation_worksheet_dataframe,
            manifest_type,
        )

        # pandas_writer.save()

    return bytesIO
