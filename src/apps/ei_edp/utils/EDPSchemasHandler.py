from common.utils.logger import Logger
import pandas as pd
import os
from django.conf import settings
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle,PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.datavalidation import DataValidation 
from common.dal.profile_da import Profile
from src.apps.copo_single_cell_submission.utils.SingleCellSchemasHandler import SingleCellSchemasHandler, SinglecellschemasSpreadsheet
from .sapio.sapio_datamanager  import Sapio
from common.utils.helpers import notify_singlecell_status, get_datetime
from common.schema_versions.lookup import dtol_lookups as lookup


class EDPSchemasHandler(SingleCellSchemasHandler):
    def write_manifest(self, profile_id, singlecell_schema, checklist_id=None, singlecell=None, file_path=None, format="xlsx", request=None):
        schema_name = singlecell_schema["name"]
        schemas = singlecell_schema["schemas"]
        checklists = singlecell_schema["checklists"]
        # component_names = singlecell_schema["components"]

        profile = Profile().get_record(profile_id)
        if not profile:
            raise Exception(f"Profile {profile_id} not found")

        # Cell formats
        alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        font = Font(bold=True)
        bolder = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        mandatory_style = NamedStyle(name="mandatory_style")
        mandatory_style.font = font
        mandatory_style.fill = PatternFill(fgColor="A6C875", fill_type = "solid")
        mandatory_style.alignment = alignment
        mandatory_style.border = bolder
        optional_style = NamedStyle(name="optional_style")
        optional_style.font = font
        optional_style.fill = PatternFill(fgColor="90D5FF", fill_type = "solid")
        optional_style.alignment = alignment
        optional_style.border = bolder
        protected_style = NamedStyle(name="protected_style")
        protected_style.font = font
        protected_style.fill = PatternFill(fgColor="D3D3D3", fill_type = "solid")
        protected_style.alignment = alignment
        protected_style.border = bolder
        component_title_style = NamedStyle(name="component_title_style")
        component_title_style.font = Font(bold=True, size=16)
        component_title_style.alignment = Alignment(horizontal="center", vertical="center")
        component_title_style.border = bolder
        component_title_style.fill = PatternFill(fgColor="FFBF00", fill_type = "solid")

        version = settings.MANIFEST_VERSION.get(schema_name, str())
        if version:
            version = "_v" + version

        workbook = load_workbook(settings.SINGLE_CELL_SCHEMAS_TEMPLATE_URL["EI_EDP"])
        worksheet_sample = workbook["sample"]
        worksheet_sample["F5"] = profile.get("jira_ticket_number", "")
        worksheet_sample["C4"] = profile.get("budget_user", "")
        worksheet_sample["C5"] = profile.get("sapio_project_id", "")

        worksheet_helper = workbook["How to complete the Manifest"]


        

        for checklist in checklists.keys():
            if checklist_id and checklist_id != checklist:
                continue

            schema_checklist = checklist
            if not checklist_id or file_path is None:
                file_path = os.path.join(
                    settings.MANIFEST_PATH,
                    settings.MANIFEST_FILE_NAME.format(
                        schema_name + "_" + checklist, version
                    ),
                )

            worksheet_helper_current_row = 4

            for component_name, schema in schemas.items():

                component_data_df = pd.DataFrame()

                if singlecell is not None:
                    component_data_df = pd.DataFrame.from_records(
                        singlecell["components"].get(component_name, [])
                    )

                worksheet_helper_current_row += 1
                
                component_schema_df = pd.DataFrame.from_records(schema)
                component_schema_df = component_schema_df.drop(
                    component_schema_df[
                        pd.isna(component_schema_df[schema_checklist])
                    ].index
                )
                if component_schema_df.empty:
                    continue

                component_schema_df.fillna("", inplace=True)
                component_schema_df["choice"] = component_schema_df[
                    component_schema_df["term_type"].isin(["enum", "suggested_enum"])
                ]["term_name"].apply(
                    lambda x: singlecell_schema.get("enums", []).get(x, [])
                )

                cell_component_title = worksheet_helper.cell(row=worksheet_helper_current_row, column=1)
                cell_component_title.value = component_name.upper() + " DETAILS"
                cell_component_title.style = component_title_style
                cell_component_title_blank = worksheet_helper.cell(row=worksheet_helper_current_row, column=2)
                cell_component_title_blank.border = bolder

                worksheet_helper_current_row += 1

                for _, field in component_schema_df.iterrows():
                    if field["term_manifest_behavior"] == "protected":
                        continue

                    label = field["term_label"]
                    description = field.get("term_description", label)

                    cell = worksheet_helper.cell(row=worksheet_helper_current_row, column=1)
                    cell.value = label
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                    if field[schema_checklist] == "M":
                        cell.style = mandatory_style 
                    elif field[schema_checklist] == "O":
                        cell.style = optional_style
                    if field["term_manifest_behavior"] == "protected":
                        cell.style = protected_style
                    
                    cell_desc = worksheet_helper.cell(row=worksheet_helper_current_row, column=2)
                    cell_desc.value = description
                    cell_desc.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                    cell_desc.border = bolder
                    
                    worksheet_helper_current_row += 1


                if component_name == "study":
                    continue

                title_row = 1
                if component_name == "sample":
                    worksheet = workbook[component_name]
                    title_row = 28
                else:
                    worksheet = workbook.create_sheet(component_name)
                    title_row = 1 
   
                column_index = 0
                for _, field in component_schema_df.iterrows():
                    column_index += 1
                    name = field["term_label"]
                    cell = worksheet.cell(row=title_row, column=column_index)
                    cell.value = name
                    description = field.get("term_description", name)
                    if description:
                        comment = Comment(field.get("term_description", name), "COPO")
                        cell.comment = comment
                    if field[checklist] == "M":    
                        cell.style = mandatory_style                                            
                    elif field[checklist] == "O":
                        cell.style = optional_style
                    if field["term_manifest_behavior"] == "protected":
                        cell.style = protected_style
                        if title_row > 1:
                            cell_highlight = worksheet.cell(row=title_row-1, column=column_index)
                            cell_highlight.fill = protected_style.fill

                    elif field["term_manifest_behavior"] == "hidden":
                        worksheet.column_dimensions[get_column_letter(column_index)].hidden = True

                    worksheet.column_dimensions[get_column_letter(column_index)].width = 17.83

                    type = field.get("term_type", "string")
                    if type in ["enum", "suggested_enum"]:
                        # Create a data-validation object with list validation
                        options = field["choice"]
                        if len(options) > 0:
                            dv = DataValidation(
                                type="list",
                                formula1='"' + ",".join(options) + '"',
                                allow_blank=True,
                                #prompt = description if description else "",
                            )
                            # Add the data-validation object to the worksheet
                            cell_start_end = '%s%d:%s%d' % (
                                get_column_letter(column_index),
                                title_row + 1,
                                get_column_letter(column_index),
                                1000,
                            )
                            worksheet.add_data_validation(dv)
                            dv.add(cell_start_end)
                    elif type == "string":
                        length = field["term_length"]
                        if length and int(length) > 0:
                            cell_start_end = '%s%d:%s%d' % (
                                get_column_letter(column_index),
                                title_row + 1,
                                get_column_letter(column_index),
                                1000,
                            )
                            dv = DataValidation(
                                type="textLength",
                                operator="lessThanOrEqual",
                                formula1=str(length),
                                allow_blank=True,
                                prompt="Please input the value with length less than or equal to " + str(length),
                                #TBC stop on error

                            )
                            # Add the data-validation object to the worksheet
                            worksheet.add_data_validation(dv)
                            dv.add(cell_start_end)
                        
                    data_row_index = title_row   
                    for _, row in component_data_df.iterrows():
                        data_row_index += 1
                        cell = worksheet.cell(row=data_row_index, column=column_index)
                        cell.value = row.get(field["term_name"], "")


        workbook.save(file_path)


class EDPSchemasSpreadsheet(SinglecellschemasSpreadsheet):
    def loadManifest(self, m_format):
        self.data = {}
        if self.profile_id is not None:
            notify_singlecell_status(
                data={"profile_id": self.profile_id},
                msg="Loading...",
                action="info",
                html_id=self.component_info,
                checklist_id=self.checklist_id,
            )
            workbook = load_workbook(self.file, data_only=True)
            profile = Profile().get_record(self.profile_id)
            sample_ws = workbook["sample"]
            sapio_project_id = str(sample_ws["C5"].value).strip()
            emails = str(sample_ws["E6"].value).strip()
            is_returned = sample_ws["L8"].value
            health_safety = sample_ws["L9"].value
            
            if not sapio_project_id or sapio_project_id != profile.get("sapio_project_id",""):
                notify_singlecell_status(
                    data={"profile_id": self.profile_id},
                    msg="Incorrect project ID: "  + str(sapio_project_id) if sapio_project_id else "Missing project ID in the manifest!",
                    action="error",
                    html_id=self.component_info,
                    checklist_id=self.checklist_id,
                )
                return False
            
            study_dict = [ {
                "study_id": profile["sapio_project_id"] if "sapio_project_id" in profile else "",
                "customer_email": emails,
                "sample_return": is_returned,
                "health_safety": health_safety}
            ]
            study_df = pd.DataFrame.from_records(study_dict)
            self.data["study"] = study_df
            self.new_data["study"] = study_df

            for sheetname in workbook.sheetnames:
                if sheetname == "How to complete the Manifest":
                    continue
                if sheetname not in self.schemas:
                    notify_singlecell_status(
                        data={"profile_id": self.profile_id},
                        msg="Incorrect worksheet in the file: "  + sheetname ,
                        action="error",
                        html_id=self.component_info,
                        checklist_id=self.checklist_id,
                    )
                    return False                
                component_ws = workbook[sheetname]
                if sheetname == "sample":  
                    component_data = self._get_ws_rows(component_ws, min_row=28)
                else:
                    component_data = self._get_ws_rows(component_ws, min_row=1)
                cols = next(component_data)
                component_data = list(component_data)
                component_df = pd.DataFrame( component_data,columns=cols)
                new_column_name = {
                    item["term_label"]: key
                    for key, item in self.schemas[sheetname].items()
                }
                component_df["Project ID"] = profile["sapio_project_id"] if "sapio_project_id" in profile else ""    
                component_df.fillna("", inplace=True)
                self.data[sheetname] = component_df
                self.new_data[sheetname] = component_df.rename(columns=new_column_name)
            """
            if "index" in workbook.sheetnames:
                index_ws = workbook["index"]
                index_data = self._get_ws_rows(index_ws, min_row=1)
                cols = next(index_data)
                index_data = list(index_data)
                index_df = pd.DataFrame(index_data,columns=cols)
                new_column_name = {
                    item["term_label"]: key
                    for key, item in self.schemas["index"].items()
                }
                index_df["Project ID"]= profile["sapio_project_id"] if "sapio_project_id" in profile else ""
                index_df.fillna("", inplace=True)
                self.data["index"] = index_df
                self.new_data["index"] = index_df.rename(columns=new_column_name)
            """
        return True
    
    def _get_ws_rows(self, worksheet, min_row=1):
        for row in worksheet.iter_rows(min_row=min_row, values_only=True):
            yield row