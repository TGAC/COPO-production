from bson import json_util
from common.dal import cursor_to_list
from common.dal.sample_da import Sample, Source
from datetime import datetime, timezone
from django.conf import settings as settings
from django.core.management import BaseCommand
from openpyxl import load_workbook
from xlrd import open_workbook, XLRDError
import ast
import importlib
import json
import pandas as pd
import re

# schema_version_path_dtol_lookups = f'web.apps.web_copo.schema_versions.{settings.CURRENT_SCHEMA_VERSION}.lookup.dtol_lookups'
schema_version_path_dtol_lookups = 'common.lookup.dtol_lookups'
dtol_lookups_data = importlib.import_module(schema_version_path_dtol_lookups)
DTOL_ENA_MAPPINGS = dtol_lookups_data.DTOL_ENA_MAPPINGS
TOL_PROFILE_TYPES = dtol_lookups_data.TOL_PROFILE_TYPES


class Command(BaseCommand):
    # The following information is shown when a user types "help"
    help = "Extract/parse \"SPECIMEN_ID\" from an .xlsx file then, retrieve metadata from COPO database based on the " \
           "data and output the result in  a .json format"

    def __init__(self):
        super().__init__()

    def add_arguments(self, parser):
        parser.add_argument("xlsx", type=str)

    def parse_db_data_to_json(self, data_in_db, json_filename):
        datetime_fields = ["date_created", "date_modified", "time_created", "time_updated"]
        df_list = []
        data_in_db = json.loads(json_util.dumps(data_in_db))

        for index in range(len(data_in_db)):
            ''' 
                Remove nested dicitonaries from the sample dictionary by retrieving
                the value of the nested dictionary and assigning it to the key of the outer dictionary
                if a dictionary is present within the list of values
            '''
            data = data_in_db[index][0]

            for key, value in data.items():
                if type(value) is dict and key == "_id":
                    field_value = value.get('$oid')
                    data[key] = field_value

                if type(value) is dict and key in datetime_fields:
                    field_value = value.get('$date')

                    # Convert datetime from milliseconds to timestamp
                    field_value = datetime.fromtimestamp(field_value / 1000.0, tz=timezone.utc).strftime(
                        '%Y-%m-%d %H:%M:%S.%f')
                    data[key] = field_value

            # Create a dictionary that maps the COPO database field name (i.e. old key)
            # to the ENA field name (i.e. new key)
            copo_and_ena_field_names_dict = {key: value['ena'] for key, value in DTOL_ENA_MAPPINGS.items() if
                                             key in list(data.keys())}

            # Replace each COPO database field name with the corresponding ENA field name
            # in the sample dictionary
            for key, value in list(data.items()):
                data[copo_and_ena_field_names_dict.get(key, key)] = data.pop(key)

            df = pd.DataFrame(data=[list(data.values())], columns=list(data.keys()))
            df_list.append(df)

        concat_df = pd.concat(df_list, ignore_index=True)  # Concatenate all the dataframes
        concat_df.to_json(json_filename)

    # A command must define handle()
    def handle(self, *args, **options):
        file_path_dict = ast.literal_eval(str(options))
        excel_file_path = file_path_dict.get("xlsx")
        try:
            open_workbook(excel_file_path)
            workbook_obj = load_workbook(filename=excel_file_path)
            worksheet = workbook_obj['nhmdump']
            specimen_id_list = []
            excel_data_dict = {}

            pattern1 = "EMu/\d{9}"
            pattern2 = "EMu/NHMUK\d{9}"  # "SPECIMEN_ID" field begins with the prefix - "NHMUK"
            pattern3 = "EMu/NHMUK\d{8}\("

            print("Commence parsing SPECIMEN_IDs from spreadsheet...")

            # Convert Excel data into a dictionary
            for key, *values in worksheet.iter_rows():
                excel_data_dict[key.value] = [v.value for v in values]

            list_of_rows = list(excel_data_dict.keys())

            # Verify that the number of rows in the spreadsheet
            # is equal to the number of items in the list
            assert len(list_of_rows) == 1853

            # Iterate through each row in the spreadsheet to retrieve the "SPECIMEN_ID"
            for row in list_of_rows:
                search_query1 = re.search(pattern1, row)
                search_query2 = re.search(pattern2, row)
                search_query3 = re.search(pattern3, row)

                if search_query1:
                    specimen_id = search_query1.group(0)[4:]
                    specimen_id_list.append(specimen_id)

                elif search_query2:
                    specimen_id = search_query2.group(0)[9:]
                    specimen_id_list.append(specimen_id)

                elif search_query3:
                    specimen_id = search_query3.group(0)[9:-1]
                    specimen_id_list.append(specimen_id)

            # Remove duplicate entries from the list
            specimen_id_list = list(set(specimen_id_list))
            assert len(specimen_id_list) == 729

            # Add the "SPECIMEN_ID" prefix to each item in the list
            # specimen_id_list = list(map(lambda x: "NHMUK" + x, specimen_id_list))
            # print('With "SPECIMEN_ID" prefix: ', specimen_id_list)

            print("Finished parsing SPECIMEN_IDs from spreadsheet...")
            print("Number of unique SPECIMEN_IDs parsed: ", len(specimen_id_list))

            specimen_ids_not_in_db = []
            samples_only_in_db = []
            sources_only_in_db = []

            # specimen_id_list = ["014421571", "014536813", "014561337", "ERGA_FP_8329_002", "EDTOLQ0405"]

            for specimen in specimen_id_list:
                print("Iterating through the list of specimen IDs...")

                sample_in_db = cursor_to_list(Sample().get_sample_by_specimen_id_regex(specimen))
                source_in_db = Source().get_by_specimen_id_regex(specimen)

                if not sample_in_db and not source_in_db:
                    specimen_ids_not_in_db.append(specimen)
                elif sample_in_db:
                    samples_only_in_db.append(sample_in_db)

                elif source_in_db:
                    sources_only_in_db.append(source_in_db)

            # Convert list of "SPECIMEN_ID" not present in COPO to json format
            if specimen_ids_not_in_db:
                print("Processing specimen_ids that are not found in the database...")
                specimen_ids_df = pd.DataFrame(data=[[specimen_ids_not_in_db]], columns=["SPECIMEN_ID"])
                specimen_ids_df.to_json('specimen_ids_not_present_in_copo_from_nhmdump_excel.json')

            # Parse db list of dictionary data to json
            if samples_only_in_db:
                # Samples only
                print("Processing samples...")
                print(f"Samples count: {len(samples_only_in_db)}")
                self.parse_db_data_to_json(samples_only_in_db, 'copo_biosamples_from_nhmdump_spreadsheet.json')

            if sources_only_in_db:
                # Sources only
                print("Processing sources...")
                print(f"Sources count: {len(sources_only_in_db)}")
                self.parse_db_data_to_json(sources_only_in_db, 'copo_biospecimens_from_nhmdump_spreadsheet.json')

            if not samples_only_in_db and not sources_only_in_db:
                print("*********************************")
                print("Error: Data in the .xlsx file do not correspond to any data in the database!")
                print("*********************************")

            print("SPECIMEN_IDs not found in db: ", specimen_ids_not_in_db)
            print("\n*********************************\n")

        except XLRDError as error:
            print("Error: ", error)
