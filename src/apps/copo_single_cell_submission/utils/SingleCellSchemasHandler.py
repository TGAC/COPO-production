from lxml import etree as ET
from .da import SinglecellSchemas
from common.utils.logger import Logger
import pandas as pd
import os
from django.conf import settings
from openpyxl.utils.cell import get_column_letter
from common.utils.helpers import notify_singlecell_status, get_datetime
from django_tools.middlewares import ThreadLocal
import inspect
import math
from common.schema_versions.lookup import dtol_lookups as lookup
from .validator import SingleCellSchemaValidators  as required_validators
from .validator import SingleCellOverallValidators as overall_validators
from common.validators.validator import Validator
from openpyxl.utils import get_column_letter
from common.dal.profile_da import Profile
import re
import collections
import numpy as np
import json
from io import BytesIO
from pyld import jsonld
from django.urls import reverse
 
l = Logger()

class SingleCellSchemasHandler:
    #def __init__(self):
        #self.headers = {'Accept': 'application/xsls' }

    def _validate_regrex(self, regex):
        if pd.isna(regex):
            return True
        
        try:
            re.compile(regex)
            return True
        except re.error:
            return False

    def _loadSchemas(self, url):
        
        xls = pd.ExcelFile(url)
        return xls

    def _parseSchemas(self, schema_name, xls):
        enums_df = pd.read_excel(xls, "allowed_values", dtype=str)
        components_df = pd.read_excel(xls, "components", index_col="key")
        standards_df = pd.read_excel(xls, "standards", index_col="key")
        technology_df = pd.read_excel(xls, "technologies", index_col="key")
        term_mapping_df = pd.read_excel(xls, "term_mapping", index_col="copo_name")
        checklist_df = pd.read_excel(xls, "checklists", index_col="key")
        schemas_df = pd.read_excel(xls, "data")

        enums_dict = {}
        for c in enums_df.columns:
            a = enums_df[c].dropna()
            enums_dict[c] = a.to_list()

        component_schemas_dict = {}
        #validate the schema
        #no empty component_name, term_name, term_label, term_type
        #no duplicate name,label within a component with same version and term_name
        #check foreign key constraints
        #check valid regex
        
        schemas_df["regex_valid"] = schemas_df["term_regex"].apply(lambda x: self._validate_regrex(x))
        if not schemas_df["regex_valid"].all():
            for c, row in schemas_df[schemas_df["regex_valid"]==False].iterrows():
                l.error(f"Invalid regex: {row['term_regex']} for {row['term_name']}")
            raise Exception("Invalid regex")

        schemas_df.drop(columns=["regex_valid"], inplace=True)

        empty_values = ""
        column_names = ["component_name", "term_name", "term_label", "term_type"]
        for i,j in zip(*np.where(pd.isnull(schemas_df[column_names]))):
            empty_values = empty_values + f"Empty value in '{column_names[j]}' at row {i+2} \n"
        if empty_values:
            raise Exception(f'Empty values: {empty_values} in the schema. Please check the schema file.')

        for checklist_id in checklist_df.index:
                #no duplicate name,label within a component with same versio and term_name
                checklist_schema_df = schemas_df.drop(schemas_df[pd.isna(schemas_df[checklist_id])].index)
                checklist_schema_df.reset_index(inplace=True)
                for c in checklist_schema_df.groupby(["component_name","term_name"]).size().reset_index().values:
                    if c[2] > 1:
                        raise Exception(f'Duplicate term_name: "{c[1]}" within a component: "{c[0]}"  with same version: {checklist_id}.')

                for c in checklist_schema_df.groupby(["component_name","term_label"]).size().reset_index().values:
                    if c[2] > 1:
                        raise Exception(f'Duplicate item_label: "{c[1]}" within a component: "{c[0]}"  with same version: {checklist_id}.')
                    
                #check foreign key constraints
                identifier_map = {}
                referenced_df = checklist_schema_df.drop(checklist_schema_df[pd.isna(checklist_schema_df["referenced_component"])].index)

                for c, component_schema_df in checklist_schema_df.groupby(["component_name"]):
                    identifier_df = component_schema_df.loc[component_schema_df['identifier'], 'term_name']                           
                    if not identifier_df.empty:
                        identifier_map[c[0]]= identifier_df.iloc[0]
                    else:
                        l.log(f"Identifier not found for {c[0]}")

                for c, row in referenced_df.iterrows():
                    if row["referenced_component"] and  row["referenced_component"] not in identifier_map.keys():
                        raise Exception(f"Referenced component: '{row['referenced_component']}' is missing")    
                    
        for c, component_schemas_df in schemas_df.groupby("component_name", sort=False):
            #component_schemas_df = schemas_df[schemas_df['component_name']== c]
            component_schemas_dict[c] = component_schemas_df.to_dict("records")

        '''
        components = schemas_df['component_name'].unique()
        for c in components:
            component_schemas_df = schemas_df[schemas_df['component_name']== c]
            component_schemas_dict[c] = component_schemas_df.to_dict("records")
        '''
        
        singlecell_dict = {}
        singlecell_dict["schemas"] =  component_schemas_dict
        singlecell_dict["enums"] = enums_dict
        singlecell_dict["standards"] = standards_df.to_dict("index")
        singlecell_dict["technologies"]= technology_df.to_dict("index")
        singlecell_dict["components"]= components_df.to_dict("index")
        singlecell_dict["checklists"]= checklist_df.to_dict("index")
        singlecell_dict["term_mapping"] = term_mapping_df.to_dict("index")
        singlecell_dict["name"] = schema_name
        singlecell_dict["deleted"] = 0
        singlecell_dict["date_modified"] = get_datetime()

        return singlecell_dict


    def write_manifest(self, singlecell_schema, checklist_id=None, singlecell=None, file_path=None, format="xlsx", request=None):
            
            schema_name = singlecell_schema["name"]
            schemas = singlecell_schema["schemas"]
            checklists = singlecell_schema["checklists"]
            #component_names = singlecell_schema["components"]

            # Cell formats
            unlocked_format = {'locked': False}

            title_format = {
            'bold' : True
            }

            desc_eg_format = {
                'text_wrap': True, 
                'italic': True,
                'font_color': '#808080'
            }

            seperator_format = {
                'bold': True,
                'align': 'left',
                'valign': 'vcenter',
                'bg_color': '#D3D3D3'
            }

            version = settings.MANIFEST_VERSION.get(schema_name, str())
            if version:
                version = "_v" + version

            for checklist in checklists.keys():
                if checklist_id and checklist_id != checklist:
                    continue
 
                schema_checklist = checklist

                if not checklist_id or file_path is None: 
                    file_path = os.path.join(settings.MANIFEST_PATH, settings.MANIFEST_FILE_NAME.format(schema_name +"_" + checklist, version)  )            
 
                data_validation_column_index = 0

                if format == "xlsx":
                    with pd.ExcelWriter(path=file_path, engine='xlsxwriter' ) as writer:  

                        for component_name, schema in schemas.items():
                            component_schema_df = pd.DataFrame.from_records(schema)
                            component_schema_df = component_schema_df.drop(component_schema_df[pd.isna(component_schema_df[schema_checklist])].index)
                            
                            if component_schema_df.empty:
                                continue
                            
                            component_schema_df.fillna("", inplace=True)
                            component_schema_df["choice"] = component_schema_df[component_schema_df["term_type"] == "enum"]["term_name"].apply(lambda x:singlecell_schema.get("enums",[]).get(x, []))
                            component_schema_df["mandatory"] = component_schema_df[schema_checklist]
                            component_schema_df.set_index(keys="term_name", inplace=True)
                
                            #component_schema_df.sort_values(by=['mandatory','term_label'], inplace=True)
                            component_schema_df.loc[component_schema_df["mandatory"] == "M" , "term_label"] = component_schema_df["term_label"]
                            component_schema_df.loc[component_schema_df["mandatory"] != "M", "term_label"] = component_schema_df["term_label"] + " (optional)"
                            component_schema_df["empty_column"] = ""

                            component_schema_df_transposed = component_schema_df.transpose()
                            component_schema_df_transposed = component_schema_df_transposed.loc[["term_label", "term_description", "term_example", "empty_column"]]
                            component_schema_df_transposed.columns = component_schema_df_transposed.iloc[0]

                            if singlecell is not None:
                                component_data_df = pd.DataFrame.from_records(singlecell["components"].get(component_name, []))
                                if not component_data_df.empty:
                                    new_column_name = { field["term_name"] : field["term_label" ]+ (" (optional)" if  field[schema_checklist] != 'M' else "") for field in schema if field[schema_checklist] in ["M","O"] }
                                    component_data_df.drop(columns=list(set(component_data_df.columns) - set(new_column_name.keys())), inplace=True)
                                    component_data_df.rename(columns=new_column_name, inplace=True)

                                    component_schema_df_transposed  = pd.concat([component_schema_df_transposed , component_data_df], axis=0)
                                    component_schema_df_transposed.fillna("", inplace=True)
                            


                            #sheet_name = component_names.get(component_name, {}).get("name", component_name)
                            #sheet_name = sheet_name[:31]
                            sheet_name = component_name
                            component_schema_df_transposed.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                            last_column_letter = get_column_letter(len(component_schema_df_transposed.columns))
                            
                            for index, field in component_schema_df.iterrows():
                                name = field["term_label"]
                                description = field.get("term_description", name)
                                type = field.get("term_type","string")
                                if name not in component_schema_df_transposed.columns:
                                    continue
                                column_index = component_schema_df_transposed.columns.get_loc(name)

                                column_length = len(description)
                                column_length = 70 if column_length > 70 else column_length
                                column_letter = get_column_letter(column_index + 1)

                                cell_format = writer.book.add_format({'num_format': '@', 'text_wrap': True, "valign":"top"})  #dosen't work
                                writer.sheets[sheet_name].set_column(column_index, column_index, column_length, cell_format)

                                if field["mandatory"] == "M":
                                    cell_format = writer.book.add_format(title_format)
                                    writer.sheets[sheet_name].conditional_format(f'{column_letter}1', {'type': 'no_errors', 'format': cell_format})                            
                                    #writer.sheets[sheet_name].write(f"{column_letter}1", name, cell_format)
                            

                                cell_start_end = '%s5:%s1048576' % (column_letter, column_letter)

                                if type == "string":
                                    pass

                                    """ doesn't work 
                                    if field["term_regex"]: 
                                        #for i in range(5, 1000):
                                        writer.sheets[sheet_name].data_validation(cell_start_end, {'validate': 'custom', 
                                                                                                            'value': f'=REGEXTEST({column_letter}5, "{field["term_regex"]}")',
                                                                                                            'error_title': 'Invalid value',
                                                                                                            'error_message': f'Invalid value for {name}. {field["term_error_message"]}',
                                                                                                            'ignore_blank': True})
                                    """
                                elif type == "enum" and "choice" in field:
                                    choice = field["choice"]

                                    if len(choice) > 0:
                                        source = ""
                                        number_of_char_for_choice = sum([len(str(x)) for x in choice])
                                        if number_of_char_for_choice <= 255:
                                            source = choice
                                        else:
                                            s = pd.Series(choice, name=field["term_label"])
                                            s.to_frame().to_excel(writer, sheet_name="data_values", index=False, header=True, startrow=0, startcol=data_validation_column_index)
                                            column_letter = get_column_letter(data_validation_column_index + 1)
                                            column_length = max(s.astype(str).map(len).max(), len(field["term_label"]))
                                            writer.sheets["data_values"].set_column(data_validation_column_index, data_validation_column_index, column_length)
                                            source = "=%s!$%s$2:$%s$%s" % ("data_values", column_letter, column_letter, str(len(choice) + 1))
                                            data_validation_column_index = data_validation_column_index + 1

                                        writer.sheets[sheet_name].data_validation(cell_start_end,
                                                                                {'validate': 'list',
                                                                                'source': source})
                                        


                            cell_format = writer.book.add_format(seperator_format)
                            writer.sheets[sheet_name].write("A4", "FILL OUT INFORMATION BELOW THIS LINE", cell_format)

                            # Set the conditional format for locking rows 2 to 3
                            cell_format = writer.book.add_format(desc_eg_format)
                            writer.sheets[sheet_name].conditional_format(f'A2:{last_column_letter}3', {'type': 'no_errors', 'format': cell_format})
                    
                            # Set all rows below row 4 to unlocked
                            cell_format = writer.book.add_format(unlocked_format)
                            for row in range(4, 1004):
                                writer.sheets[sheet_name].set_row(row, None, cell_format)
                            
                            # Protect the worksheet
                            writer.sheets[sheet_name].protect()
                        
                    if "data_values" in writer.sheets:
                        writer.sheets["data_values"].protect()
                        writer.sheets["data_values"].hide()

                    """
                    for sheet in writer.sheets.values():
                        sheet.autofit()
                    """
    
                elif format == "jsonld":

                     
                    url_prefix = settings.COPO_URL.get(settings.ENVIRONMENT_TYPE, "local")
                    singlecell_dict = {}

                    context = []
                    
                    for component_name, schema in schemas.items():
                        component_schema_df = pd.DataFrame.from_records(schema)
                        component_schema_df = component_schema_df.drop(component_schema_df[pd.isna(component_schema_df[schema_checklist])].index)
                        component_schema_df.fillna("", inplace=True)

                        if component_schema_df.empty:
                            continue


                    for component_name, schema in schemas.items():
                        
                        if not schema :
                            continue
                        
                        if singlecell is not None:
                            component_data_df = pd.DataFrame.from_records(singlecell["components"].get(component_name, []))
                            
                            if not component_data_df.empty:
                                context.append(request.build_absolute_uri(f"/media/assets/manifests/{schema_name}_{checklist}_{component_name}{version}.jsonld"))
                                new_column_name = {field["term_name"] : url_prefix + reverse("copo_single_cell_submission:display_term", args=[singlecell_schema["name"], field["term_name"]]) if not field["term_reference"] or not isinstance(field["term_reference"],str) or field["term_type"] == "ontology" else field["term_reference"] for field in schema if field[schema_checklist] in ["M","O"] }

                                component_data_df.drop(columns=list(set(component_data_df.columns) -set(new_column_name.keys())), inplace=True)
                                component_data_df.rename(columns=new_column_name, inplace=True)
                                component_data_df.fillna("", inplace=True)
                                component_data_df["@type"] = url_prefix + reverse("copo_single_cell_submission:display_term", args=[singlecell_schema["name"], component_name])
                                singlecell_dict[url_prefix + reverse("copo_single_cell_submission:display_term", args=[singlecell_schema["name"], component_name])] = component_data_df.to_dict("records")
             
                    try:
                        compacted = jsonld.compact(singlecell_dict, context)
                    except Exception as e:
                        l.error(f"Failed to generate JSON-LD for {schema_name} {checklist}: {str(e)}")
                        l.exception(e)
                        continue
    
                    if isinstance(file_path, BytesIO):
                        outfile = file_path
                    else:
                        outfile = open(file_path, "w")
                    #with outfile:
                    outfile.write(bytes(json.dumps(compacted), 'utf-8'))


    def updateSchemas(self):
        for name, url in settings.SINGLE_CELL_SCHEMAS_URL.items():
            version = settings.MANIFEST_VERSION.get(name, str())
            #url = f"singlecell_schema_main_v{version}.xlsx"
            xls = self._loadSchemas(url)
            singlecell_schema = self._parseSchemas(name,xls)
            singlecell_schema["version"] = version
            SinglecellSchemas().get_collection_handle().find_one_and_update({"name": singlecell_schema["name"]},
                                                                                {"$set": singlecell_schema},
                                                                                upsert=True)
            self.write_manifest(singlecell_schema)
            self.generate_jsonld(singlecell_schema, None)
         

    def generate_jsonld(self, singlecell_schema, checklist_id=None):
        """
        Generate JSON-LD for a single cell study.
        The JSON-LD will be generated based on the schema and the single cell schema.
        """
        #schemas = SinglecellSchemas().get_schema(schema_name=schema_name, target_id=checklist_id)
        #singlecell_schemas = SinglecellSchemas().get_collection_handle().find_one({"name":schema_name})

        url_prefix = settings.COPO_URL.get(settings.ENVIRONMENT_TYPE, "local")
        checklists = singlecell_schema.get("checklists", {})

        schema_name = singlecell_schema.get("name", "COPO_SINGLE_CELL")
        version = settings.MANIFEST_VERSION.get(schema_name, str())
        if version:
            version = "_v" + version

        for id, checklist in checklists.items():
            if checklist_id and id != checklist_id:
                continue

            schemas = singlecell_schema.get("schemas",{})

            for component_name, schema in schemas.items():
                component_schema_df = pd.DataFrame.from_records(schema)
                component_schema_df = component_schema_df.drop(component_schema_df[pd.isna(component_schema_df[id])].index)
                if component_schema_df.empty:
                    continue
                component_schema_df.fillna("", inplace=True)
                file_path = os.path.join(settings.MANIFEST_PATH, settings.MANIFEST_JSONLD_FILE_NAME.format(schema_name, id, component_name, version) )            

                context_for_checklist_component = {}
                #context_for_checklist_component[component_name]={"@id": f"{ei_url}/{component_name}", "@container": "@set"}
                context_for_checklist_component[component_name]={"@id": url_prefix + reverse("copo_single_cell_submission:display_term", args=[singlecell_schema["name"], component_name]), "@container": "@set"}
                

                for _, row in component_schema_df.iterrows():
                    term_reference = row["term_reference"]
                    term_type = row["term_type"]
                    term_name = row["term_name"]

                    if term_reference and term_type != "ontology":
                        pass
                    elif (not term_reference or term_type == "ontology"):
                        term_reference =  url_prefix + reverse("copo_single_cell_submission:display_term", args=[singlecell_schema["name"], term_name])   
                    else:
                        continue

                    context_for_checklist_component[term_name] = {"@id": term_reference}
                
                try:
                    compacted = jsonld.compact({}, context_for_checklist_component)
                except Exception as e:
                    l.error(f"Failed to generate JSON-LD context for {schema_name} {id} {component_name}: {str(e)}")
                    continue
                                                
                with open(file_path, "w") as outfile:
                    outfile.write(json.dumps(compacted))

class SinglecellschemasSpreadsheet:
   def __init__(self, file, profile_id, schema_name, checklist_id,  component, validators=[]):
        self.req = ThreadLocal.get_current_request()
        self.profile_id = profile_id
        self.checklist_id = checklist_id
        self.schema_name = schema_name
        self.data = {}
        self.new_data = {}
        self.component = component
        self.component_info = f"{self.component}_info"
        self.component_table = f"{self.component}_table"
        self.required_validators = []    
        self.overall_validators = []
        self.symbiont_list = []
        self.validator_list = []
        self.schemas = None
        self.filenames = []
        # if a file is passed in, then this is the first time we have seen the spreadsheet,
        # if not then we are looking at creating samples having previously validated
        if file:
            self.file = file
        #else:
        #    self.sample_data = self.req.session.get( self.component_table, "")
        #    self.isupdate = self.req.session.get("isupdate", False)
        # create list of required validators
    
        required = dict(globals().items())["required_validators"]
        for element_name in dir(required):
            element = getattr(required, element_name)
            if inspect.isclass(element) and issubclass(element, Validator) and not element.__name__ == "Validator":
                self.required_validators.append(element)

        required = dict(globals().items())["overall_validators"]
        for element_name in dir(required):
            element = getattr(required, element_name)
            if inspect.isclass(element) and issubclass(element, Validator) and not element.__name__ == "Validator":
                self.overall_validators.append(element)
     

   def get_filenames_from_manifest(self):
        #return list(self.data["File name"])
        file_map = {}
        msg = ""
        for component, df in self.new_data.items():
            schema = self.schemas[component]        
            schema_df = pd.DataFrame.from_dict(list(schema.values()))
            schema_file_df = schema_df.loc[schema_df['term_type'] == 'file', ["term_name"]]
            if schema_file_df.empty:
                continue

            schema_file_checksum_df = schema_df.loc[schema_df['term_name'].isin([name+"_checksum" for name in schema_file_df["term_name"].tolist()]), ["term_name"]]
            if not schema_file_checksum_df.empty:
                schema_file_checksum_df["term_checksum"] = schema_file_checksum_df["term_name"]
                schema_file_checksum_df["term_name"] = schema_file_checksum_df["term_name"].str.replace("_checksum", "", regex=False)
                schema_file_df = schema_file_df.merge(schema_file_checksum_df, on="term_name", how="left")
                schema_file_df.fillna("", inplace=True)
            else:
                schema_file_df["term_checksum"] = ""

            for index, row in schema_file_df.iterrows():
                if row["term_checksum"]:
                    df2 = df[[row["term_name"], row["term_checksum"]]]  
                else:
                    df2 = df[[row["term_name"]]]    
                df2 = df2.dropna()
                for index2, row2 in df2.iterrows():
                    file_name = row2[row["term_name"]].strip()
                    if file_name:
                        if row["term_checksum"]:
                            if row2[row["term_checksum"]].strip():
                                file_map[row2[row["term_name"]]] = row2[row["term_checksum"]]
                            else:
                                msg = msg + "File name: " + file_name + f' is missing checksum <b>{row["term_checksum"]}</b> in the manifest.<br/>'
                                
                        else:
                            file_map[file_name] = ""
     
        filelist =  list(file_map.keys()) 

        #find duplicated files
        duplicated_files = [item for item, count in collections.Counter(filelist).items() if count > 1]
        if duplicated_files:
            msg = msg + "Duplicated files: " + ", ".join(duplicated_files) + " in the manifest.<br/>"
            
        if msg: 
            return False, msg
        
        self.filename_map = file_map            
        return file_map, msg

   def loadManifest(self, m_format):

        if self.profile_id is not None:
            notify_singlecell_status(data={"profile_id": self.profile_id}, msg="Loading...", action="info",
                            html_id=self.component_info, checklist_id=self.checklist_id)

            try:
                # read excel and convert all to string
                if m_format == "xls":
                    self.data = pd.read_excel(self.file, keep_default_na=False,
                                                  na_values=lookup.NA_VALS, sheet_name=None, dtype=str)    

                    if not isinstance(self.data, dict): 
                        raise Exception("invalid single cell manifest") 
                         
                    #remove the data_values sheet which is used for dropdowns
                    self.data.pop("data_values", None)

                else:
                    raise Exception("Unknown file format")
                
                #if self.data.empty:
                #    raise Exception("Empty file")
                #profile = Profile().get_record(self.profile_id)
                #schema_name = profile.get("schema_name", "COPO_SINGLE_CELL")

                for key, df in self.data.items():  
                    #df = df.iloc[3:]  # remove the first 3 rows                  
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                    df = df.apply(lambda x: x.astype(str))
                    df = df.apply(lambda x: x.str.strip())
                    #self.data.columns = self.data.columns.str.replace(" ", "")
                   
                singlecell = SinglecellSchemas().get_collection_handle().find_one({"name":self.schema_name},{"schemas":1, "enums":1})
                self.schemas = singlecell["schemas"]

                if self.schemas:
                    for component in list(self.schemas.keys()):    
                        component_schema_df = pd.DataFrame.from_records(self.schemas[component])
                        component_schema_df = component_schema_df.drop(component_schema_df[pd.isna(component_schema_df[self.checklist_id])].index)
                        if component_schema_df.empty:
                            self.schemas.pop(component, None)
                            continue
                        component_schema_df.fillna("", inplace=True)
                        component_schema_df["choice"] = component_schema_df[component_schema_df["term_type"] == "enum"]["term_name"].apply(lambda x:singlecell["enums"].get(x, []))
                        component_schema_df["mandatory"] = component_schema_df[self.checklist_id]
                        component_schema_df.set_index(keys="term_name", inplace=True)

                        self.schemas[component] = component_schema_df.to_dict("index")

                    for component, df in self.data.items():
                        if not component in self.schemas.keys():
                            raise Exception("Invalid worksheet: " + component)
                        new_column_name = { name : name.replace(" (optional)", "",-1) for name in df.columns.values.tolist() }
                        self.new_data[component] = df.rename(columns=new_column_name)    
                        new_column_name = {item["term_label"] : key for key, item in self.schemas[component].items() }
                        self.new_data[component].rename(columns=new_column_name, inplace=True)
                        self.new_data[component] = self.new_data[component].iloc[3:]  # remove the first 3 rows

            except Exception as e:
                # if error notify via web socket
                l.exception(e)
                notify_singlecell_status(data={"profile_id": self.profile_id}, msg="Unable to load file. " + str(e),
                                action="error",
                                html_id=self.component_info, checklist_id=self.checklist_id)
                return False
            return True

   def validate(self):
        flag = True
        errors = []
        warnings = []
        self.isupdate = False
 
        # checklist = EnaChecklist().get_checklist(self.checklist_id)  
        # validate for required fields

        for component, df in self.new_data.items():
            for v in self.required_validators:
                try:
                    errors, warnings, flag, self.isupdate = v(profile_id=self.profile_id, schema=self.schemas[component], component=component,
                                                            data=df, fields=None,
                                                            errors=errors, warnings=warnings, flag=flag,
                                                            isupdate=self.isupdate).validate()
                except Exception as e:
                    l.exception(e)
                    error_message = str(e).replace("<", "").replace(">", "")

                    flag = False
                    errors.append(error_message)
                                        
        for v in self.overall_validators:
            try:
                errors, warnings, flag, self.isupdate = v(profile_id=self.profile_id, schemas=self.schemas,
                                                        data=self.new_data, fields=None,
                                                        errors=errors, warnings=warnings, flag=flag,
                                                        isupdate=self.isupdate).validate()
            except Exception as e:
                l.exception(e)
                error_message = str(e).replace("<", "").replace(">", "")

                flag = False
                errors.append(error_message)


        # send warnings
        if warnings:
            l.log(",".join(warnings))
            notify_singlecell_status(data={"profile_id": self.profile_id},
                            msg="<br>".join(warnings),
                            action="warning",
                            html_id="warning_info2", checklist_id=self.checklist_id)
        # if flag is false, compile list of errors
        if not flag:
            errors = list(map(lambda x: "<li>" + x + "</li>", errors))
            errors = "".join(errors)
            l.log(errors)
            notify_singlecell_status(data={"profile_id": self.profile_id},
                            msg="<h4>" + self.file.name + "</h4><h2>Errors</h2><ol>" + errors + "</ol>",
                            action="error",
                            html_id=self.component_info, checklist_id=self.checklist_id)
            return False, errors
 
        for component, df in self.new_data.items():
            for column in df.columns:
                if column.startswith(Validator.PREFIX_4_NEW_FIELD):
                    self.data[column.removeprefix(Validator.PREFIX_4_NEW_FIELD)] = self.new_data[column]

        # if we get here we have a valid spreadsheet
        #notify_singlecell_status(data={"profile_id": self.profile_id}, msg="Spreadsheet is valid", action="info",
        #                html_id=self.component_info)

        return True, errors

   def collect(self):
        # create table data to show to the frontend from parsed manifest
        singlecell_data = {}

        for component, df in self.data.items():
            df = df.iloc[3:]  # remove the first 3 rows
            singlecell_data[component] = []
            headers = list()
            for col in list(df.columns):
                headers.append(col)
            singlecell_data[component].append(headers)
            for index, row in df.iterrows():
                r = list(row)
                for idx, x in enumerate(r):
                    if x is math.nan:
                        r[idx] = ""
                singlecell_data[component].append(r)
        # store sample data in the session to be used to create mongo objects
        self.req.session[f"{self.component}_data"] = singlecell_data
        self.req.session["checklist_id"] = self.checklist_id
        self.req.session["filename_map"] = self.filename_map

        notify_singlecell_status(data={"profile_id": self.profile_id, "components": list(self.data.keys())}, msg=singlecell_data, action="make_table",
                        html_id=f"{self.component}_parse_table", checklist_id=self.checklist_id)
        
